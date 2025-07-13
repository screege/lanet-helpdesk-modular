#!/usr/bin/env python3
"""
Verificar que el archivo Excel se puede leer correctamente
"""

import os
from openpyxl import load_workbook

def verify_excel_file():
    """Verify the Excel file can be opened"""
    print("🔍 VERIFYING EXCEL FILE")
    print("=" * 50)
    
    # Find the most recent report file
    reports_dir = os.path.join(os.path.dirname(__file__), '..', 'reports_files')
    reports_dir = os.path.abspath(reports_dir)

    # Also check current directory reports_files
    if not os.path.exists(reports_dir):
        reports_dir = os.path.join(os.getcwd(), 'reports_files')
        reports_dir = os.path.abspath(reports_dir)
    
    print(f"📁 Reports directory: {reports_dir}")
    
    if not os.path.exists(reports_dir):
        print("❌ Reports directory doesn't exist")
        return
    
    # Find Excel files
    excel_files = [f for f in os.listdir(reports_dir) if f.endswith('.xlsx')]
    
    if not excel_files:
        print("❌ No Excel files found")
        return
    
    # Get the most recent file
    excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(reports_dir, x)), reverse=True)
    latest_file = excel_files[0]
    file_path = os.path.join(reports_dir, latest_file)
    
    print(f"📄 Testing file: {latest_file}")
    print(f"📁 Full path: {file_path}")
    
    try:
        # Try to open the file
        wb = load_workbook(file_path)
        print("✅ File opened successfully with openpyxl")
        
        # Get the active sheet
        ws = wb.active
        print(f"📊 Sheet name: {ws.title}")
        print(f"📊 Max row: {ws.max_row}")
        print(f"📊 Max column: {ws.max_column}")
        
        # Read first few rows
        print("\n📋 First 5 rows:")
        for row in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):  # First 10 columns
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    cell_value = ""
                row_data.append(str(cell_value)[:30])  # Limit to 30 chars
            print(f"Row {row}: {' | '.join(row_data)}")
        
        # Check for problematic characters
        print("\n🔍 Checking for problematic characters...")
        problem_found = False
        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # Check for control characters
                    for char in cell_value:
                        if ord(char) < 32 and char not in '\t\n\r':
                            print(f"⚠️ Found control character (ord {ord(char)}) in row {row}, col {col}")
                            problem_found = True
                            break
        
        if not problem_found:
            print("✅ No problematic characters found")
        
        wb.close()
        print("\n✅ File verification completed successfully!")
        
    except Exception as e:
        print(f"❌ Error opening file: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")

if __name__ == "__main__":
    verify_excel_file()
