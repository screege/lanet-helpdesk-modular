#!/usr/bin/env python3
"""
Test specific Excel file
"""

import os
from openpyxl import load_workbook

def test_specific_file():
    """Test specific Excel file"""
    print("🔍 TESTING SPECIFIC EXCEL FILE")
    print("=" * 50)
    
    file_path = os.path.join(os.getcwd(), 'reports_files', 'reporte_consolidado_2025_07.xlsx')
    
    print(f"📄 Testing file: {file_path}")
    print(f"📁 File exists: {os.path.exists(file_path)}")
    
    if not os.path.exists(file_path):
        print("❌ File doesn't exist")
        return
    
    file_size = os.path.getsize(file_path)
    print(f"📁 File size: {file_size} bytes")
    
    try:
        # Try to open the file
        wb = load_workbook(file_path)
        print("✅ File opened successfully with openpyxl")
        
        # Get the active sheet
        ws = wb.active
        print(f"📊 Sheet name: {ws.title}")
        print(f"📊 Max row: {ws.max_row}")
        print(f"📊 Max column: {ws.max_column}")
        
        # Read headers
        print("\n📋 Headers:")
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=7, column=col).value  # Headers are in row 7
            headers.append(str(header) if header else "")
        print(" | ".join(headers))
        
        # Read first few data rows
        print("\n📋 First 3 data rows:")
        for row in range(8, min(11, ws.max_row + 1)):  # Data starts at row 8
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    cell_value = ""
                # Limit display to 20 chars per cell
                row_data.append(str(cell_value)[:20])
            print(f"Row {row-7}: {' | '.join(row_data)}")
        
        wb.close()
        print("\n✅ File verification completed successfully!")
        print("🎉 The Excel file should open correctly in Microsoft Excel now!")
        
    except Exception as e:
        print(f"❌ Error opening file: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")

if __name__ == "__main__":
    test_specific_file()
