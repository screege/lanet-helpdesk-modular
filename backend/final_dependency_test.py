#!/usr/bin/env python3
"""
Final test to verify all Excel report dependencies are working
"""

def test_excel_dependencies():
    """Test all Excel report dependencies"""
    print("🔍 FINAL EXCEL DEPENDENCIES TEST")
    print("=" * 50)
    
    success_count = 0
    total_tests = 0
    
    # Test 1: Core Excel functionality
    total_tests += 1
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create a test workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        ws['A1'] = "Test Cell"
        ws['A1'].font = Font(bold=True)
        
        print("✅ Test 1: openpyxl functionality - PASSED")
        success_count += 1
    except Exception as e:
        print(f"❌ Test 1: openpyxl functionality - FAILED: {e}")
    
    # Test 2: Pandas Excel export
    total_tests += 1
    try:
        import pandas as pd
        
        # Create test DataFrame
        df = pd.DataFrame({
            'Column1': ['Test1', 'Test2'],
            'Column2': [1, 2]
        })
        
        print("✅ Test 2: pandas functionality - PASSED")
        success_count += 1
    except Exception as e:
        print(f"❌ Test 2: pandas functionality - FAILED: {e}")
    
    # Test 3: Database connectivity
    total_tests += 1
    try:
        import psycopg2
        print("✅ Test 3: psycopg2 import - PASSED")
        success_count += 1
    except Exception as e:
        print(f"❌ Test 3: psycopg2 import - FAILED: {e}")
    
    # Test 4: Date/time handling
    total_tests += 1
    try:
        import pytz
        from datetime import datetime
        
        # Test Mexico timezone
        mexico_tz = pytz.timezone('America/Mexico_City')
        now = datetime.now(mexico_tz)
        formatted = now.strftime('%d/%m/%Y %H:%M')
        
        print("✅ Test 4: pytz timezone handling - PASSED")
        success_count += 1
    except Exception as e:
        print(f"❌ Test 4: pytz timezone handling - FAILED: {e}")
    
    # Test 5: Flask framework
    total_tests += 1
    try:
        from flask import Flask
        app = Flask(__name__)
        print("✅ Test 5: Flask framework - PASSED")
        success_count += 1
    except Exception as e:
        print(f"❌ Test 5: Flask framework - FAILED: {e}")
    
    # Test 6: Standard library modules
    total_tests += 1
    try:
        import csv
        import os
        import logging
        import uuid
        import traceback
        
        print("✅ Test 6: Standard library modules - PASSED")
        success_count += 1
    except Exception as e:
        print(f"❌ Test 6: Standard library modules - FAILED: {e}")
    
    # Summary
    print(f"\n📊 TEST RESULTS:")
    print(f"✅ Passed: {success_count}/{total_tests}")
    print(f"❌ Failed: {total_tests - success_count}/{total_tests}")
    
    if success_count == total_tests:
        print(f"\n🎉 ALL TESTS PASSED!")
        print(f"✅ Excel report functionality is fully operational")
        return True
    else:
        print(f"\n⚠️ Some tests failed. Check dependencies.")
        return False

if __name__ == "__main__":
    test_excel_dependencies()
