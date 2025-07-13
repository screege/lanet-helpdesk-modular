#!/usr/bin/env python3
"""
Find the exact source of Excel corruption
"""

import psycopg2
import re

def find_problematic_characters():
    """Find exactly what's causing Excel corruption"""
    print("🔍 FINDING EXCEL CORRUPTION SOURCE")
    print("=" * 60)
    
    try:
        # Connect to database
        conn = psycopg2.connect(
            host="localhost",
            database="lanet_helpdesk",
            user="postgres",
            password="Poikl55+*"
        )
        cursor = conn.cursor()
        
        # Get actual data that's being used in reports
        query = """
        SELECT 
            t.ticket_number,
            c.name as client_name,
            t.subject,
            t.status::text as status,
            t.priority::text as priority,
            t.created_at,
            t.resolved_at,
            u.name as technician_name,
            s.name as site_name,
            'Pendiente' as resolution_notes
        FROM tickets t
        LEFT JOIN clients c ON t.client_id = c.client_id
        LEFT JOIN users u ON t.assigned_to = u.user_id
        LEFT JOIN sites s ON t.site_id = s.site_id
        WHERE t.created_at >= '2025-07-01' AND t.created_at <= '2025-07-31'
        ORDER BY t.created_at DESC
        """
        
        cursor.execute(query)
        tickets = cursor.fetchall()
        
        print(f"📋 Analyzing {len(tickets)} tickets for problematic characters...")
        
        problematic_chars = set()
        problematic_records = []
        
        for i, ticket in enumerate(tickets):
            fields = [
                ('ticket_number', ticket[0]),
                ('client_name', ticket[1]),
                ('subject', ticket[2]),
                ('status', ticket[3]),
                ('priority', ticket[4]),
                ('technician_name', ticket[7]),
                ('site_name', ticket[8]),
                ('resolution_notes', ticket[9])
            ]
            
            for field_name, value in fields:
                if value is not None:
                    str_value = str(value)
                    
                    # Check each character
                    for char_pos, char in enumerate(str_value):
                        char_code = ord(char)
                        
                        # Flag problematic characters
                        is_problematic = False
                        reason = ""
                        
                        # Control characters (except tab, newline, carriage return)
                        if char_code < 32 and char not in '\t\n\r':
                            is_problematic = True
                            reason = "control_character"
                        
                        # High Unicode characters that might cause XML issues
                        elif char_code > 127 and char not in 'áéíóúñüÁÉÍÓÚÑÜ¿¡':
                            is_problematic = True
                            reason = "high_unicode"
                        
                        # Characters that might be interpreted as formulas
                        elif char_pos == 0 and char in '=+-@':
                            is_problematic = True
                            reason = "formula_start"
                        
                        # XML-problematic characters
                        elif char in '<>&"\'':
                            is_problematic = True
                            reason = "xml_special"
                        
                        if is_problematic:
                            problematic_chars.add((char, char_code, reason))
                            problematic_records.append({
                                'ticket': ticket[0],
                                'field': field_name,
                                'char': char,
                                'char_code': char_code,
                                'reason': reason,
                                'position': char_pos,
                                'context': str_value[max(0, char_pos-5):char_pos+6]
                            })
        
        # Report findings
        if problematic_chars:
            print(f"\n⚠️ Found {len(problematic_chars)} types of problematic characters:")
            for char, code, reason in sorted(problematic_chars):
                print(f"  '{char}' (code {code}) - {reason}")
            
            print(f"\n📋 First 10 problematic instances:")
            for record in problematic_records[:10]:
                print(f"  Ticket {record['ticket']}, field {record['field']}")
                print(f"    Char: '{record['char']}' (code {record['char_code']}) - {record['reason']}")
                print(f"    Context: ...{record['context']}...")
                print()
        else:
            print("✅ No obviously problematic characters found")
        
        # Now create a SUPER CLEAN version
        print(f"\n🧹 Creating SUPER CLEAN test file...")
        create_super_clean_excel(tickets)
        
        cursor.close()
        conn.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        print(traceback.format_exc())

def super_clean_string(value):
    """Ultra-aggressive string cleaning"""
    if value is None:
        return ""
    
    str_value = str(value).strip()
    if not str_value:
        return ""
    
    # Keep ONLY basic ASCII letters, numbers, spaces, and basic punctuation
    allowed = set('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 .,()-')
    
    cleaned = ''.join(char if char in allowed else ' ' for char in str_value)
    
    # Remove multiple spaces
    cleaned = ' '.join(cleaned.split())
    
    # Limit length aggressively
    if len(cleaned) > 50:
        cleaned = cleaned[:50]
    
    # Ensure it doesn't start with formula characters
    if cleaned and cleaned[0] in '=+-@':
        cleaned = ' ' + cleaned
    
    return cleaned

def create_super_clean_excel(tickets):
    """Create Excel with super clean data"""
    try:
        import pandas as pd
        import os
        
        # Prepare super clean data
        clean_data = []
        
        for ticket in tickets[:20]:  # Test with first 20 tickets only
            clean_row = {
                'Numero': super_clean_string(ticket[0]),
                'Cliente': super_clean_string(ticket[1]),
                'Asunto': super_clean_string(ticket[2]),
                'Estado': super_clean_string(ticket[3]),
                'Prioridad': super_clean_string(ticket[4]),
                'Creado': ticket[5].strftime('%d/%m/%Y') if ticket[5] else 'N/A',
                'Resuelto': ticket[6].strftime('%d/%m/%Y') if ticket[6] else 'Pendiente',
                'Tecnico': super_clean_string(ticket[7]),
                'Sitio': super_clean_string(ticket[8]),
                'Solucion': super_clean_string(ticket[9])
            }
            clean_data.append(clean_row)
        
        # Create DataFrame
        df = pd.DataFrame(clean_data)
        
        # Save to Excel
        file_path = os.path.join(os.getcwd(), 'reports_files', 'super_clean_test.xlsx')
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Use the most basic Excel writer settings
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Reporte', index=False)
        
        print(f"✅ Super clean test file created: {file_path}")
        
        # Verify it can be opened
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        print(f"✅ File verification: Opens successfully with openpyxl")
        print(f"📊 Rows: {wb.active.max_row}, Columns: {wb.active.max_column}")
        wb.close()
        
        return file_path
        
    except Exception as e:
        print(f"❌ Error creating super clean Excel: {e}")
        import traceback
        print(traceback.format_exc())
        return None

if __name__ == "__main__":
    find_problematic_characters()
