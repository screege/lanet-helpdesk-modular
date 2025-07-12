#!/usr/bin/env python3
"""
Test minimal Excel generation with clean data
"""

import os
import pandas as pd
from datetime import datetime

def create_minimal_excel():
    """Create minimal Excel file with clean test data"""
    print("🔍 CREATING MINIMAL EXCEL TEST")
    print("=" * 50)
    
    try:
        # Create minimal test data
        test_data = [
            {
                'Número': 'TKT-000001',
                'Cliente': 'Test Client',
                'Asunto': 'Test Subject',
                'Estado': 'abierto',
                'Prioridad': 'media',
                'Creado': '01/07/2025 10:00',
                'Resuelto': 'Pendiente',
                'Técnico': 'Test Tech',
                'Sitio': 'Test Site',
                'Solución': 'Test Solution'
            },
            {
                'Número': 'TKT-000002',
                'Cliente': 'Another Client',
                'Asunto': 'Another Subject',
                'Estado': 'cerrado',
                'Prioridad': 'alta',
                'Creado': '02/07/2025 11:00',
                'Resuelto': '02/07/2025 15:00',
                'Técnico': 'Another Tech',
                'Sitio': 'Another Site',
                'Solución': 'Problem resolved'
            }
        ]
        
        # Create DataFrame
        df = pd.DataFrame(test_data)
        
        # File path
        file_path = os.path.join(os.getcwd(), 'reports_files', 'test_minimal.xlsx')
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        print(f"📄 Creating test file: {file_path}")
        
        # Create Excel file with minimal options
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Test', index=False, startrow=5, header=True)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Test']
            
            # Add simple headers
            worksheet['A1'] = 'LANET SYSTEMS - TEST REPORT'
            worksheet['A2'] = 'Test Period: 01/07/2025 - 31/07/2025'
            worksheet['A3'] = 'Generated: ' + datetime.now().strftime('%d/%m/%Y %H:%M')
            worksheet['A4'] = 'Total Records: 2'
        
        print(f"✅ Test Excel file created: {file_path}")
        
        # Verify file can be opened
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        print(f"✅ File verification: Can be opened with openpyxl")
        print(f"📊 Sheets: {wb.sheetnames}")
        ws = wb.active
        print(f"📊 Max row: {ws.max_row}, Max column: {ws.max_column}")
        wb.close()
        
        return file_path
        
    except Exception as e:
        print(f"❌ Error creating minimal Excel: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return None

def create_excel_with_real_data():
    """Create Excel with actual database data but ultra-clean"""
    print("\n🔍 CREATING EXCEL WITH REAL DATA")
    print("=" * 50)
    
    try:
        # Import database connection
        import psycopg2
        
        # Connect to database
        conn = psycopg2.connect(
            host="localhost",
            database="lanet_helpdesk",
            user="postgres",
            password="Poikl55+*"
        )
        cursor = conn.cursor()
        
        # Get a few tickets with simple query - BASIC COLUMNS ONLY
        query = """
        SELECT
            t.ticket_number,
            COALESCE(c.name, 'Sin cliente') as client_name,
            COALESCE(t.subject, 'Sin asunto') as subject,
            t.status::text as status,
            t.priority::text as priority,
            t.created_at,
            t.resolved_at,
            COALESCE(u.name, 'Sin asignar') as technician_name,
            COALESCE(s.name, 'Sin sitio') as site_name,
            'Pendiente' as resolution_notes
        FROM tickets t
        LEFT JOIN clients c ON t.client_id = c.client_id
        LEFT JOIN users u ON t.assigned_to = u.user_id
        LEFT JOIN sites s ON t.site_id = s.site_id
        WHERE t.created_at >= '2025-07-01' AND t.created_at <= '2025-07-31'
        ORDER BY t.created_at DESC
        LIMIT 10
        """
        
        cursor.execute(query)
        tickets = cursor.fetchall()
        
        print(f"📋 Found {len(tickets)} tickets")
        
        # Clean data aggressively
        clean_data = []
        for ticket in tickets:
            # Ultra-aggressive cleaning
            clean_row = {
                'Número': str(ticket[0]) if ticket[0] else 'N/A',
                'Cliente': clean_string(ticket[1]),
                'Asunto': clean_string(ticket[2]),
                'Estado': clean_string(ticket[3]),
                'Prioridad': clean_string(ticket[4]),
                'Creado': ticket[5].strftime('%d/%m/%Y %H:%M') if ticket[5] else 'N/A',
                'Resuelto': ticket[6].strftime('%d/%m/%Y %H:%M') if ticket[6] else 'Pendiente',
                'Técnico': clean_string(ticket[7]),
                'Sitio': clean_string(ticket[8]),
                'Solución': clean_string(ticket[9]) if ticket[9] else 'Pendiente'
            }
            clean_data.append(clean_row)
        
        # Create DataFrame
        df = pd.DataFrame(clean_data)
        
        # File path
        file_path = os.path.join(os.getcwd(), 'reports_files', 'test_real_data.xlsx')
        
        print(f"📄 Creating real data file: {file_path}")
        
        # Create Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Reporte Consolidado', index=False, startrow=5, header=True)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Reporte Consolidado']
            
            # Add headers
            worksheet['A1'] = 'LANET SYSTEMS - REPORTE CONSOLIDADO DE SOPORTE TÉCNICO'
            worksheet['A2'] = 'REPORTE CONSOLIDADO - TODOS LOS CLIENTES'
            worksheet['A3'] = 'Período: 01/07/2025 - 31/07/2025'
            worksheet['A4'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            worksheet['A5'] = f"Total de Tickets: {len(clean_data)}"
        
        print(f"✅ Real data Excel file created: {file_path}")
        
        # Verify
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        print(f"✅ File verification: Can be opened with openpyxl")
        wb.close()
        
        cursor.close()
        conn.close()
        
        return file_path
        
    except Exception as e:
        print(f"❌ Error creating real data Excel: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return None

def clean_string(value):
    """Ultra-clean string for Excel"""
    if value is None:
        return ""
    
    # Convert to string
    str_value = str(value).strip()
    
    if not str_value:
        return ""
    
    # Keep only safe ASCII and basic Spanish characters
    safe_chars = []
    for char in str_value:
        char_code = ord(char)
        if (32 <= char_code <= 126 or 
            char in 'áéíóúñüÁÉÍÓÚÑÜ'):
            safe_chars.append(char)
        elif char in '\t\n\r':
            safe_chars.append(' ')
    
    # Join and clean
    cleaned = ''.join(safe_chars)
    cleaned = ' '.join(cleaned.split())
    
    # Limit length
    if len(cleaned) > 100:
        cleaned = cleaned[:100]
    
    return cleaned if cleaned else ""

if __name__ == "__main__":
    # Test minimal first
    minimal_file = create_minimal_excel()
    
    # Test with real data
    if minimal_file:
        real_file = create_excel_with_real_data()
        
        if real_file:
            print(f"\n🎉 SUCCESS! Both files created successfully")
            print(f"📄 Minimal test: {minimal_file}")
            print(f"📄 Real data test: {real_file}")
        else:
            print(f"\n⚠️ Minimal worked, but real data failed")
    else:
        print(f"\n❌ Even minimal test failed")
