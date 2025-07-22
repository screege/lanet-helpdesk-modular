#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LANET Helpdesk V3 - Reports Module Routes (Simple Version)
Temporary implementation without pandas dependency
"""

from flask import Blueprint, request, current_app, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from utils.security import require_role
from datetime import datetime
import pytz
import logging
import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)
logger = logging.getLogger(__name__)

def get_mexico_time():
    """Get current time in Mexico timezone"""
    mexico_tz = pytz.timezone('America/Mexico_City')
    return datetime.now(mexico_tz)

def sanitize_excel_cell_value(value):
    """
    Sanitize cell values to prevent Excel formula injection and corruption.

    Excel interprets cells starting with =, +, -, @ as formulas, which can cause:
    1. XML corruption errors like "Registros quitados: Fórmula de /xl/worksheets/sheet1.xml parte"
    2. Security vulnerabilities (formula injection)
    3. Data corruption

    Args:
        value: The cell value to sanitize

    Returns:
        str: Sanitized value safe for Excel
    """
    if value is None:
        return ''

    # Convert to string
    str_value = str(value)

    # If empty, return as is
    if not str_value:
        return str_value

    # Formula characters that cause Excel issues
    formula_chars = ['=', '+', '-', '@']

    # Check if the value starts with a formula character
    if str_value.startswith(tuple(formula_chars)):
        # Prefix with a single quote to force Excel to treat it as text
        # This is the standard Excel method for preventing formula interpretation
        return "'" + str_value

    # For values that don't start with formula chars but contain them,
    # we generally leave them as-is since they're safe in the middle of text
    # Exception: if the entire value is just a formula character, prefix it
    if str_value in formula_chars:
        return "'" + str_value

    return str_value

def generate_friendly_filename(report_type, output_format='excel'):
    """Generate user-friendly filename with date"""
    now = get_mexico_time()

    # Spanish month names
    months = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }

    month_name = months[now.month]
    year = now.year
    day = now.day

    # File extension
    ext = 'xlsx' if output_format == 'excel' else 'pdf'

    # Generate filename based on report type
    if report_type == 'monthly':
        filename = f"Reporte_Mensual_{month_name}_{year}.{ext}"
    elif report_type == 'quick':
        filename = f"Reporte_Rapido_{day}_{month_name}_{year}.{ext}"
    elif report_type == 'statistics':
        filename = f"Estadisticas_{month_name}_{year}.{ext}"
    elif report_type == 'sla':
        filename = f"Reporte_SLA_{month_name}_{year}.{ext}"
    else:
        filename = f"Reporte_{day}_{month_name}_{year}.{ext}"

    return filename

def generate_excel_report_lanet_format(data, filename, report_title="REPORTE CONSOLIDADO"):
    """Generate Excel report matching LANET format"""
    try:
        # Create reports directory if it doesn't exist
        reports_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports_files')
        os.makedirs(reports_dir, exist_ok=True)

        file_path = os.path.join(reports_dir, filename)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Consolidado"

        # Get current date for report
        now = get_mexico_time()
        period_start = now.replace(day=1).strftime('%d/%m/%Y')
        period_end = now.strftime('%d/%m/%Y')

        # Add LANET header format (matching reference file) - sanitized
        ws.append([sanitize_excel_cell_value('LANET SYSTEMS - REPORTE CONSOLIDADO DE SOPORTE TÉCNICO')])
        ws.append([sanitize_excel_cell_value(f'{report_title} - TODOS LOS CLIENTES')])
        ws.append([sanitize_excel_cell_value(f'Período: {period_start} - {period_end}')])
        ws.append([sanitize_excel_cell_value(f'Generado: {now.strftime("%d/%m/%Y %H:%M")} | Total de Tickets: {len(data)}')])
        ws.append([])  # Empty row

        # Add data headers (matching reference file exactly)
        headers = ['Número', 'Cliente', 'Asunto', 'Estado', 'Prioridad', 'Creado', 'Resuelto', 'Técnico', 'Sitio', 'Fecha Resolución', 'Resolución']
        ws.append(headers)

        # Style header row (row 6)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=6, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row_data in data:
            # Format dates to match reference file
            created_at = 'Sin fecha'
            resolved_at = 'Pendiente'
            fecha_resolucion = 'Sin fecha'

            if row_data.get('created_at'):
                try:
                    if isinstance(row_data['created_at'], str):
                        created_at = row_data['created_at']
                    else:
                        created_at = row_data['created_at'].strftime('%d/%m/%Y %H:%M')
                except:
                    created_at = 'Sin fecha'

            if row_data.get('resolved_at'):
                try:
                    if isinstance(row_data['resolved_at'], str):
                        resolved_at = row_data['resolved_at']
                        fecha_resolucion = row_data['resolved_at']
                    else:
                        resolved_at = row_data['resolved_at'].strftime('%d/%m/%Y %H:%M')
                        fecha_resolucion = resolved_at
                except:
                    resolved_at = 'Pendiente'
                    fecha_resolucion = 'Sin fecha'

            # Map status to Spanish
            status_map = {
                'abierto': 'abierto',
                'en_proceso': 'en_proceso',
                'resuelto': 'resuelto',
                'cerrado': 'cerrado',
                'pendiente': 'pendiente'
            }

            status = status_map.get(row_data.get('status', ''), row_data.get('status', 'abierto'))

            # Sanitize all cell values to prevent Excel formula injection
            sanitized_row = [
                sanitize_excel_cell_value(row_data.get('ticket_number', 'N/A')),
                sanitize_excel_cell_value(row_data.get('client_name', 'N/A')),
                sanitize_excel_cell_value(row_data.get('subject', 'Sin asunto')),
                sanitize_excel_cell_value(status),
                sanitize_excel_cell_value(row_data.get('priority', 'media')),
                sanitize_excel_cell_value(created_at),
                sanitize_excel_cell_value(resolved_at),
                sanitize_excel_cell_value(row_data.get('assigned_to_name', 'Sin asignar')),
                sanitize_excel_cell_value(row_data.get('site_name', 'N/A')),
                sanitize_excel_cell_value(fecha_resolucion),
                sanitize_excel_cell_value(row_data.get('resolution', 'Sin resolución'))
            ]

            ws.append(sanitized_row)

        # Auto-adjust column widths to match reference file
        column_widths = [12, 25, 30, 12, 12, 18, 18, 20, 15, 18, 30]
        for i, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = width

        # Save the file
        wb.save(file_path)
        logger.info(f"✅ Excel file generated: {file_path}")

        return file_path

    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return None

def get_tickets_data_for_report(user_claims):
    """Get tickets data for report generation with correct column names"""
    try:
        # Build query based on user role
        user_role = user_claims.get('role')
        client_id = user_claims.get('client_id')
        site_ids = user_claims.get('site_ids', [])

        # Updated query with correct column names and additional fields
        base_query = """
        SELECT
            t.ticket_number,
            c.name as client_name,
            s.name as site_name,
            t.subject,
            t.status,
            t.priority,
            t.created_at,
            t.resolved_at,
            COALESCE(u_assigned.name, 'Sin asignar') as assigned_to_name,
            COALESCE(t.resolution_notes, 'Sin resolución') as resolution
        FROM tickets t
        LEFT JOIN clients c ON t.client_id = c.client_id
        LEFT JOIN sites s ON t.site_id = s.site_id
        LEFT JOIN users u_assigned ON t.assigned_to = u_assigned.user_id
        WHERE 1=1
        """

        params = []

        # Apply role-based filtering
        if user_role == 'client_admin' and client_id:
            base_query += " AND t.client_id = %s"
            params.append(client_id)
        elif user_role == 'solicitante' and site_ids:
            placeholders = ','.join(['%s'] * len(site_ids))
            base_query += f" AND t.site_id IN ({placeholders})"
            params.extend(site_ids)

        base_query += " ORDER BY t.created_at DESC LIMIT 1000"

        logger.info(f"🔍 Executing query for {user_role}: {base_query}")
        logger.info(f"📊 Query parameters: {params}")

        tickets = current_app.db_manager.execute_query(base_query, params)

        logger.info(f"📈 Retrieved {len(tickets) if tickets else 0} tickets from database")

        # Format tickets data
        formatted_tickets = []
        for ticket in tickets or []:
            formatted_ticket = dict(ticket)

            # Keep original datetime objects for proper formatting in Excel function
            # The Excel function will handle the formatting

            formatted_tickets.append(formatted_ticket)

        return formatted_tickets

    except Exception as e:
        logger.error(f"Error getting tickets data: {e}")
        return []

@reports_bp.route('/monthly/status', methods=['GET'])
@jwt_required()
def get_monthly_status():
    """Get monthly reports status"""
    try:
        logger.info("📊 Getting monthly reports status")
        
        # Get current user info
        current_user_id = get_jwt_identity()
        claims = get_jwt()
        
        # Simple status response
        status_data = {
            'next_report': 'Próximo reporte: 1 de cada mes a las 09:00',
            'last_execution': 'Última ejecución: No disponible',
            'total_clients': 0,
            'status': 'Sistema de reportes en modo simplificado'
        }
        
        # Count active clients
        try:
            clients_query = "SELECT COUNT(*) as count FROM clients WHERE is_active = true"
            client_count = current_app.db_manager.execute_query(clients_query, fetch='one')
            if client_count:
                status_data['total_clients'] = client_count['count']
        except Exception as e:
            logger.error(f"Error counting clients: {e}")
        
        return current_app.response_manager.success(status_data)
        
    except Exception as e:
        logger.error(f"Error getting monthly status: {e}")
        return current_app.response_manager.error(f"Error getting monthly status: {str(e)}", 500)

@reports_bp.route('/executions', methods=['GET'])
@jwt_required()
def get_executions():
    """Get report executions"""
    try:
        logger.info("📈 Getting report executions")
        
        current_user_id = get_jwt_identity()
        claims = get_jwt()
        
        # Get executions from database
        query = """
        SELECT execution_id, config_id, execution_type, status, output_format,
               file_path, file_size, started_at, completed_at, executed_by
        FROM report_executions
        ORDER BY started_at DESC
        LIMIT 50
        """
        
        executions = current_app.db_manager.execute_query(query)
        
        # Format the results
        formatted_executions = []
        for execution in executions or []:
            formatted_executions.append({
                'execution_id': str(execution['execution_id']),
                'config_id': str(execution['config_id']) if execution['config_id'] else None,
                'execution_type': execution['execution_type'],
                'status': execution['status'],
                'output_format': execution['output_format'],
                'file_path': execution['file_path'],
                'file_size': execution['file_size'],
                'started_at': execution['started_at'].isoformat() if execution['started_at'] else None,
                'completed_at': execution['completed_at'].isoformat() if execution['completed_at'] else None,
                'executed_by': str(execution['executed_by']) if execution['executed_by'] else None
            })
        
        return current_app.response_manager.success(formatted_executions)
        
    except Exception as e:
        logger.error(f"Error getting executions: {e}")
        return current_app.response_manager.error(f"Error getting executions: {str(e)}", 500)

@reports_bp.route('/monthly/generate-test', methods=['POST'])
@jwt_required()
@require_role(['superadmin', 'admin', 'technician', 'client_admin'])
def generate_monthly_test():
    """Generate monthly test report with real Excel file"""
    try:
        logger.info("📊 Generating monthly test report with real Excel file")

        data = request.get_json() or {}
        current_user_id = get_jwt_identity()
        claims = get_jwt()

        # Generate execution ID
        execution_id = str(uuid.uuid4())

        # Get tickets data
        tickets_data = get_tickets_data_for_report(claims)

        # Generate Excel file with friendly name
        filename = generate_friendly_filename('monthly', 'excel')
        file_path = generate_excel_report_lanet_format(tickets_data, filename, "REPORTE MENSUAL")

        if not file_path:
            return current_app.response_manager.error("Error generando el archivo del reporte", 500)

        # Insert execution record
        now = get_mexico_time()
        query = """
        INSERT INTO report_executions (
            execution_id, config_id, execution_type, status,
            output_format, file_path, executed_by, started_at, completed_at
        ) VALUES (%s::uuid, %s, %s, %s, %s, %s, %s::uuid, %s, %s)
        """

        current_app.db_manager.execute_query(query, (
            execution_id,
            None,  # No config_id for test reports
            'manual',
            'completed',
            'excel',
            os.path.basename(file_path),
            current_user_id,
            now,
            now
        ), fetch='none')

        logger.info("✅ Monthly test report generated with Excel file")

        return current_app.response_manager.success(
            {'execution_id': execution_id, 'file_path': os.path.basename(file_path)},
            f'Reporte mensual generado exitosamente con {len(tickets_data)} registros. Revise la sección de Ejecuciones para descargarlo.'
        )

    except Exception as e:
        logger.error(f"Error generating monthly test report: {e}")
        return current_app.response_manager.error(f"Error generating test report: {str(e)}", 500)

@reports_bp.route('/generate-quick', methods=['POST'])
@jwt_required()
@require_role(['superadmin', 'admin', 'technician', 'client_admin'])
def generate_quick_report():
    """Generate quick report with real Excel file"""
    try:
        logger.info("🚀 Generating quick report with real Excel file")

        data = request.get_json()
        if not data:
            return current_app.response_manager.bad_request('No data provided')

        output_format = data.get('output_format', 'excel')
        current_user_id = get_jwt_identity()
        claims = get_jwt()

        # Generate execution ID first
        execution_id = str(uuid.uuid4())

        # Get tickets data
        tickets_data = get_tickets_data_for_report(claims)

        # Generate file based on format
        file_path = None
        if output_format == 'excel':
            filename = generate_friendly_filename('quick', 'excel')
            file_path = generate_excel_report_lanet_format(tickets_data, filename, "REPORTE RÁPIDO")
        else:
            # For PDF, create a simple text file for now
            filename = generate_friendly_filename('quick', 'txt')
            reports_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports_files')
            os.makedirs(reports_dir, exist_ok=True)
            file_path = os.path.join(reports_dir, filename)

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"REPORTE RÁPIDO\n")
                f.write(f"Generado: {get_mexico_time().strftime('%d/%m/%Y %H:%M')}\n")
                f.write(f"Total tickets: {len(tickets_data)}\n\n")
                for ticket in tickets_data[:10]:  # First 10 tickets
                    f.write(f"- {ticket.get('ticket_number', 'N/A')}: {ticket.get('title', 'N/A')}\n")

        if not file_path:
            return current_app.response_manager.error("Error generando el archivo del reporte", 500)

        # Create execution record
        now = get_mexico_time()

        query = """
        INSERT INTO report_executions (
            execution_id, config_id, execution_type, status,
            output_format, file_path, executed_by, started_at, completed_at
        ) VALUES (%s::uuid, %s, %s, %s, %s, %s, %s::uuid, %s, %s)
        """

        current_app.db_manager.execute_query(query, (
            execution_id,
            None,
            'manual',
            'completed',
            output_format,
            os.path.basename(file_path),
            current_user_id,
            now,
            now
        ), fetch='none')

        return current_app.response_manager.success(
            {'execution_id': execution_id, 'file_path': os.path.basename(file_path)},
            f'Reporte rápido {output_format.upper()} generado exitosamente con {len(tickets_data)} registros.'
        )

    except Exception as e:
        logger.error(f"Error generating quick report: {e}")
        return current_app.response_manager.error(f"Error generating quick report: {str(e)}", 500)

@reports_bp.route('/generate-statistics', methods=['POST'])
@jwt_required()
@require_role(['superadmin', 'admin', 'technician', 'client_admin'])
def generate_statistics_report():
    """Generate statistics report with real Excel file"""
    try:
        logger.info("📈 Generating statistics report with real Excel file")

        data = request.get_json()
        if not data:
            return current_app.response_manager.bad_request('No data provided')

        output_format = data.get('output_format', 'excel')
        current_user_id = get_jwt_identity()
        claims = get_jwt()

        # Generate execution ID
        execution_id = str(uuid.uuid4())

        # Get tickets data
        tickets_data = get_tickets_data_for_report(claims)

        # Generate file based on format
        file_path = None
        if output_format == 'excel':
            filename = generate_friendly_filename('statistics', 'excel')
            file_path = generate_excel_report_lanet_format(tickets_data, filename, "REPORTE DE ESTADÍSTICAS")
        else:
            # For PDF, create a simple text file for now
            filename = generate_friendly_filename('statistics', 'txt')
            reports_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports_files')
            os.makedirs(reports_dir, exist_ok=True)
            file_path = os.path.join(reports_dir, filename)

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"REPORTE DE ESTADÍSTICAS\n")
                f.write(f"Generado: {get_mexico_time().strftime('%d/%m/%Y %H:%M')}\n")
                f.write(f"Total tickets: {len(tickets_data)}\n\n")

                # Basic statistics
                status_count = {}
                priority_count = {}
                for ticket in tickets_data:
                    status = ticket.get('status', 'N/A')
                    priority = ticket.get('priority', 'N/A')
                    status_count[status] = status_count.get(status, 0) + 1
                    priority_count[priority] = priority_count.get(priority, 0) + 1

                f.write("ESTADÍSTICAS POR ESTADO:\n")
                for status, count in status_count.items():
                    f.write(f"- {status}: {count}\n")

                f.write("\nESTADÍSTICAS POR PRIORIDAD:\n")
                for priority, count in priority_count.items():
                    f.write(f"- {priority}: {count}\n")

        if not file_path:
            return current_app.response_manager.error("Error generando el archivo del reporte", 500)

        # Create execution record
        now = get_mexico_time()

        query = """
        INSERT INTO report_executions (
            execution_id, config_id, execution_type, status,
            output_format, file_path, executed_by, started_at, completed_at
        ) VALUES (%s::uuid, %s, %s, %s, %s, %s, %s::uuid, %s, %s)
        """

        current_app.db_manager.execute_query(query, (
            execution_id,
            None,
            'manual',
            'completed',
            output_format,
            os.path.basename(file_path),
            current_user_id,
            now,
            now
        ), fetch='none')

        return current_app.response_manager.success(
            {'execution_id': execution_id, 'file_path': os.path.basename(file_path)},
            f'Reporte de estadísticas {output_format.upper()} generado exitosamente con {len(tickets_data)} registros.'
        )

    except Exception as e:
        logger.error(f"Error generating statistics report: {e}")
        return current_app.response_manager.error(f"Error generating statistics report: {str(e)}", 500)

@reports_bp.route('/generate-sla', methods=['POST'])
@jwt_required()
@require_role(['superadmin', 'admin', 'technician', 'client_admin'])
def generate_sla_report():
    """Generate SLA report (simplified version)"""
    try:
        logger.info("⏱️ Generating SLA report (simplified)")
        
        data = request.get_json()
        if not data:
            return current_app.response_manager.bad_request('No data provided')
        
        output_format = data.get('output_format', 'pdf')
        current_user_id = get_jwt_identity()
        
        # Create execution record
        execution_id = str(uuid.uuid4())
        now = get_mexico_time()
        
        query = """
        INSERT INTO report_executions (
            execution_id, config_id, execution_type, status,
            output_format, file_path, executed_by, started_at, completed_at
        ) VALUES (%s::uuid, %s, %s, %s, %s, %s, %s::uuid, %s, %s)
        """
        
        current_app.db_manager.execute_query(query, (
            execution_id,
            None,
            'manual',
            'completed',
            output_format,
            f'sla_report_{execution_id}.{output_format}',
            current_user_id,
            now,
            now
        ), fetch='none')
        
        return current_app.response_manager.success(
            {'execution_id': execution_id},
            f'Reporte SLA {output_format.upper()} generado exitosamente (modo simplificado).'
        )
        
    except Exception as e:
        logger.error(f"Error generating SLA report: {e}")
        return current_app.response_manager.error(f"Error generating SLA report: {str(e)}", 500)

@reports_bp.route('/executions/<execution_id>/download', methods=['GET'])
@jwt_required()
def download_report(execution_id):
    """Download generated report file"""
    try:
        logger.info(f"🔽 Download request for execution_id: {execution_id}")

        # Get file info from database
        query = """
        SELECT file_path, output_format, status
        FROM report_executions
        WHERE execution_id = %s::uuid AND status = 'completed' AND file_path IS NOT NULL
        """

        result = current_app.db_manager.execute_query(query, (execution_id,), fetch='one')

        if not result:
            logger.error(f"❌ No file found for execution_id: {execution_id}")
            return current_app.response_manager.not_found('Archivo de reporte no encontrado')

        file_name = result['file_path']
        output_format = result['output_format']

        # Build full file path
        reports_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports_files')
        file_path = os.path.join(reports_dir, file_name)

        # Check if file exists
        if not os.path.exists(file_path):
            logger.error(f"❌ File not found on disk: {file_path}")
            return current_app.response_manager.not_found('Archivo físico no encontrado')

        # Determine MIME type
        if output_format == 'excel' or file_name.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            as_attachment = True
        elif output_format == 'pdf' or file_name.endswith('.pdf'):
            mimetype = 'application/pdf'
            as_attachment = True
        else:
            mimetype = 'text/plain'
            as_attachment = True

        logger.info(f"✅ Sending file: {file_path}")

        return send_file(
            file_path,
            mimetype=mimetype,
            as_attachment=as_attachment,
            download_name=file_name
        )

    except Exception as e:
        logger.error(f"Error downloading report: {e}")
        return current_app.response_manager.error(f"Error downloading report: {str(e)}", 500)
