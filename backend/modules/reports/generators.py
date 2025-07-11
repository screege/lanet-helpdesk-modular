#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LANET Helpdesk V3 - Report Generators
Professional report generation with PDF, Excel, and CSV support
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Any
import io

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Base class for report generation"""
    
    def __init__(self):
        self.reports_dir = os.path.join(os.getcwd(), 'reports')
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
    
    def generate_report(self, file_path: str, config: Dict, data: Dict, output_format: str):
        """Generate report in specified format"""
        try:
            if output_format == 'pdf':
                return self._generate_pdf(file_path, config, data)
            elif output_format == 'excel':
                return self._generate_excel(file_path, config, data)
            elif output_format == 'csv':
                return self._generate_csv(file_path, config, data)
            else:
                raise ValueError(f"Unsupported output format: {output_format}")
                
        except Exception as e:
            logger.error(f"Error generating {output_format} report: {e}")
            raise
    
    def _generate_pdf(self, file_path: str, config: Dict, data: Dict):
        """Generate PDF report using ReportLab"""
        try:
            # Try to import ReportLab, fallback to simple text if not available
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.lib import colors
                
                # Create PDF document
                doc = SimpleDocTemplate(file_path, pagesize=A4)
                styles = getSampleStyleSheet()
                story = []
                
                # Title
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    spaceAfter=30,
                    alignment=1  # Center alignment
                )

                # Get period information for title
                period_info = data.get('period', {})
                report_title = f"LANET Helpdesk - {config.get('name', 'Reporte del Sistema')}"
                if period_info.get('month_name'):
                    report_title += f" - {period_info['month_name']}"

                story.append(Paragraph(report_title, title_style))
                story.append(Spacer(1, 20))
                
                # Report info with Spanish formatting
                period_info = data.get('period', {})
                report_type_names = {
                    'dashboard': 'Resumen Ejecutivo',
                    'tickets': 'Análisis de Tickets',
                    'sla': 'Cumplimiento SLA',
                    'performance': 'Rendimiento de Técnicos'
                }

                report_info = [
                    ['Reporte:', config.get('name', 'Sin nombre')],
                    ['Tipo:', report_type_names.get(config.get('report_type', 'dashboard'), 'Resumen Ejecutivo')],
                    ['Generado:', datetime.utcnow().strftime('%d de %B de %Y a las %H:%M:%S UTC')],
                    ['Período:', f"{period_info.get('start_date_formatted', 'N/A')} - {period_info.get('end_date_formatted', 'N/A')}"]
                ]
                
                info_table = Table(report_info, colWidths=[2*inch, 4*inch])
                info_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(info_table)
                story.append(Spacer(1, 30))
                
                # Metrics sections
                if 'tickets' in data:
                    story.append(Paragraph("Métricas de Tickets", styles['Heading2']))
                    story.append(Spacer(1, 12))
                    
                    tickets_data = data['tickets']
                    ticket_metrics = [
                        ['Métrica', 'Valor'],
                        ['Total de Tickets', str(tickets_data.get('total_tickets', 0))],
                        ['Tickets Abiertos', str(tickets_data.get('open_tickets', 0))],
                        ['Tickets Resueltos', str(tickets_data.get('resolved_tickets', 0))],
                        ['Tickets Cerrados', str(tickets_data.get('closed_tickets', 0))],
                        ['Prioridad Alta', str(tickets_data.get('high_priority', 0))],
                        ['Prioridad Media', str(tickets_data.get('medium_priority', 0))],
                        ['Prioridad Baja', str(tickets_data.get('low_priority', 0))],
                        ['Tiempo Promedio Resolución (hrs)', f"{tickets_data.get('avg_resolution_hours', 0):.1f}"]
                    ]
                    
                    tickets_table = Table(ticket_metrics, colWidths=[3*inch, 2*inch])
                    tickets_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(tickets_table)
                    story.append(Spacer(1, 20))
                
                if 'sla' in data:
                    story.append(Paragraph("Métricas de SLA", styles['Heading2']))
                    story.append(Spacer(1, 12))
                    
                    sla_data = data['sla']
                    sla_metrics = [
                        ['Métrica SLA', 'Valor'],
                        ['Tickets con SLA', str(sla_data.get('total_tickets_with_sla', 0))],
                        ['Incumplimientos Respuesta', str(sla_data.get('response_breaches', 0))],
                        ['Incumplimientos Resolución', str(sla_data.get('resolution_breaches', 0))],
                        ['Cumplimiento Respuesta (%)', f"{sla_data.get('response_compliance', 0):.1f}%"],
                        ['Cumplimiento Resolución (%)', f"{sla_data.get('resolution_compliance', 0):.1f}%"],
                        ['Tiempo Promedio Respuesta (min)', f"{sla_data.get('avg_response_minutes', 0):.1f}"],
                        ['Tiempo Promedio Resolución (min)', f"{sla_data.get('avg_resolution_minutes', 0):.1f}"]
                    ]
                    
                    sla_table = Table(sla_metrics, colWidths=[3*inch, 2*inch])
                    sla_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(sla_table)
                    story.append(Spacer(1, 20))
                
                if 'technicians' in data and data['technicians'].get('technicians'):
                    story.append(Paragraph("Rendimiento de Técnicos", styles['Heading2']))
                    story.append(Spacer(1, 12))
                    
                    tech_data = [['Técnico', 'Tickets Asignados', 'Tickets Resueltos', 'Tiempo Promedio (hrs)']]
                    
                    for tech in data['technicians']['technicians'][:10]:  # Top 10
                        tech_data.append([
                            f"{tech.get('first_name', '')} {tech.get('last_name', '')}",
                            str(tech.get('assigned_tickets', 0)),
                            str(tech.get('resolved_tickets', 0)),
                            f"{tech.get('avg_resolution_hours', 0):.1f}"
                        ])
                    
                    tech_table = Table(tech_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                    tech_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(tech_table)
                
                # Build PDF
                doc.build(story)
                logger.info(f"PDF report generated successfully: {file_path}")
                
            except ImportError:
                # Fallback to simple text file if ReportLab not available
                logger.warning("ReportLab not available, generating text file instead")
                self._generate_text_fallback(file_path, config, data, 'PDF')
                
        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            raise
    
    def _generate_excel(self, file_path: str, config: Dict, data: Dict):
        """Generate Excel report using openpyxl"""
        try:
            # Try to import openpyxl, fallback to simple text if not available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                
                # Remove default sheet and create custom sheets
                wb.remove(wb.active)
                
                # Summary sheet
                ws_summary = wb.create_sheet("Resumen")
                
                # Header
                ws_summary['A1'] = "LANET Helpdesk - Reporte del Sistema"
                ws_summary['A1'].font = Font(size=16, bold=True)
                ws_summary['A1'].alignment = Alignment(horizontal='center')
                ws_summary.merge_cells('A1:D1')
                
                # Report info
                row = 3
                ws_summary[f'A{row}'] = "Reporte:"
                ws_summary[f'B{row}'] = config.get('name', 'Sin nombre')
                row += 1
                ws_summary[f'A{row}'] = "Tipo:"
                ws_summary[f'B{row}'] = config.get('report_type', 'dashboard').title()
                row += 1
                ws_summary[f'A{row}'] = "Generado:"
                ws_summary[f'B{row}'] = datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')
                row += 1
                ws_summary[f'A{row}'] = "Período:"
                ws_summary[f'B{row}'] = f"{data.get('period', {}).get('start_date', 'N/A')} - {data.get('period', {}).get('end_date', 'N/A')}"
                
                # Style header cells
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                for r in range(3, 7):
                    ws_summary[f'A{r}'].fill = header_fill
                    ws_summary[f'A{r}'].font = Font(bold=True)
                
                # Tickets data
                if 'tickets' in data:
                    ws_tickets = wb.create_sheet("Tickets")
                    tickets_data = data['tickets']
                    
                    # Headers
                    headers = ['Métrica', 'Valor']
                    for col, header in enumerate(headers, 1):
                        cell = ws_tickets.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                    
                    # Data
                    metrics = [
                        ['Total de Tickets', tickets_data.get('total_tickets', 0)],
                        ['Tickets Abiertos', tickets_data.get('open_tickets', 0)],
                        ['Tickets Resueltos', tickets_data.get('resolved_tickets', 0)],
                        ['Tickets Cerrados', tickets_data.get('closed_tickets', 0)],
                        ['Prioridad Alta', tickets_data.get('high_priority', 0)],
                        ['Prioridad Media', tickets_data.get('medium_priority', 0)],
                        ['Prioridad Baja', tickets_data.get('low_priority', 0)],
                        ['Tiempo Promedio Resolución (hrs)', f"{tickets_data.get('avg_resolution_hours', 0):.1f}"]
                    ]
                    
                    for row, (metric, value) in enumerate(metrics, 2):
                        ws_tickets.cell(row=row, column=1, value=metric)
                        ws_tickets.cell(row=row, column=2, value=value)
                
                # SLA data
                if 'sla' in data:
                    ws_sla = wb.create_sheet("SLA")
                    sla_data = data['sla']
                    
                    # Headers
                    for col, header in enumerate(['Métrica SLA', 'Valor'], 1):
                        cell = ws_sla.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                    
                    # Data
                    sla_metrics = [
                        ['Tickets con SLA', sla_data.get('total_tickets_with_sla', 0)],
                        ['Incumplimientos Respuesta', sla_data.get('response_breaches', 0)],
                        ['Incumplimientos Resolución', sla_data.get('resolution_breaches', 0)],
                        ['Cumplimiento Respuesta (%)', f"{sla_data.get('response_compliance', 0):.1f}%"],
                        ['Cumplimiento Resolución (%)', f"{sla_data.get('resolution_compliance', 0):.1f}%"],
                        ['Tiempo Promedio Respuesta (min)', f"{sla_data.get('avg_response_minutes', 0):.1f}"],
                        ['Tiempo Promedio Resolución (min)', f"{sla_data.get('avg_resolution_minutes', 0):.1f}"]
                    ]
                    
                    for row, (metric, value) in enumerate(sla_metrics, 2):
                        ws_sla.cell(row=row, column=1, value=metric)
                        ws_sla.cell(row=row, column=2, value=value)
                
                # Technicians data
                if 'technicians' in data and data['technicians'].get('technicians'):
                    ws_tech = wb.create_sheet("Técnicos")
                    
                    # Headers
                    tech_headers = ['Técnico', 'Tickets Asignados', 'Tickets Resueltos', 'Tiempo Promedio (hrs)']
                    for col, header in enumerate(tech_headers, 1):
                        cell = ws_tech.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                    
                    # Data
                    for row, tech in enumerate(data['technicians']['technicians'], 2):
                        ws_tech.cell(row=row, column=1, value=f"{tech.get('first_name', '')} {tech.get('last_name', '')}")
                        ws_tech.cell(row=row, column=2, value=tech.get('assigned_tickets', 0))
                        ws_tech.cell(row=row, column=3, value=tech.get('resolved_tickets', 0))
                        ws_tech.cell(row=row, column=4, value=f"{tech.get('avg_resolution_hours', 0):.1f}")
                
                # Auto-adjust column widths
                for ws in wb.worksheets:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save workbook
                wb.save(file_path)
                logger.info(f"Excel report generated successfully: {file_path}")
                
            except ImportError:
                # Fallback to simple text file if openpyxl not available
                logger.warning("openpyxl not available, generating text file instead")
                self._generate_text_fallback(file_path, config, data, 'Excel')
                
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise

    def _generate_csv(self, file_path: str, config: Dict, data: Dict):
        """Generate CSV report"""
        try:
            import csv

            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # Header
                writer.writerow(['LANET Helpdesk - Reporte del Sistema'])
                writer.writerow([])
                writer.writerow(['Reporte:', config.get('name', 'Sin nombre')])
                writer.writerow(['Tipo:', config.get('report_type', 'dashboard').title()])
                writer.writerow(['Generado:', datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')])
                writer.writerow(['Período:', f"{data.get('period', {}).get('start_date', 'N/A')} - {data.get('period', {}).get('end_date', 'N/A')}"])
                writer.writerow([])

                # Tickets data
                if 'tickets' in data:
                    writer.writerow(['=== MÉTRICAS DE TICKETS ==='])
                    tickets_data = data['tickets']

                    writer.writerow(['Métrica', 'Valor'])
                    writer.writerow(['Total de Tickets', tickets_data.get('total_tickets', 0)])
                    writer.writerow(['Tickets Abiertos', tickets_data.get('open_tickets', 0)])
                    writer.writerow(['Tickets Resueltos', tickets_data.get('resolved_tickets', 0)])
                    writer.writerow(['Tickets Cerrados', tickets_data.get('closed_tickets', 0)])
                    writer.writerow(['Prioridad Alta', tickets_data.get('high_priority', 0)])
                    writer.writerow(['Prioridad Media', tickets_data.get('medium_priority', 0)])
                    writer.writerow(['Prioridad Baja', tickets_data.get('low_priority', 0)])
                    writer.writerow(['Tiempo Promedio Resolución (hrs)', f"{tickets_data.get('avg_resolution_hours', 0):.1f}"])
                    writer.writerow([])

                # SLA data
                if 'sla' in data:
                    writer.writerow(['=== MÉTRICAS DE SLA ==='])
                    sla_data = data['sla']

                    writer.writerow(['Métrica SLA', 'Valor'])
                    writer.writerow(['Tickets con SLA', sla_data.get('total_tickets_with_sla', 0)])
                    writer.writerow(['Incumplimientos Respuesta', sla_data.get('response_breaches', 0)])
                    writer.writerow(['Incumplimientos Resolución', sla_data.get('resolution_breaches', 0)])
                    writer.writerow(['Cumplimiento Respuesta (%)', f"{sla_data.get('response_compliance', 0):.1f}%"])
                    writer.writerow(['Cumplimiento Resolución (%)', f"{sla_data.get('resolution_compliance', 0):.1f}%"])
                    writer.writerow(['Tiempo Promedio Respuesta (min)', f"{sla_data.get('avg_response_minutes', 0):.1f}"])
                    writer.writerow(['Tiempo Promedio Resolución (min)', f"{sla_data.get('avg_resolution_minutes', 0):.1f}"])
                    writer.writerow([])

                # Technicians data
                if 'technicians' in data and data['technicians'].get('technicians'):
                    writer.writerow(['=== RENDIMIENTO DE TÉCNICOS ==='])
                    writer.writerow(['Técnico', 'Tickets Asignados', 'Tickets Resueltos', 'Tiempo Promedio (hrs)'])

                    for tech in data['technicians']['technicians']:
                        writer.writerow([
                            f"{tech.get('first_name', '')} {tech.get('last_name', '')}",
                            tech.get('assigned_tickets', 0),
                            tech.get('resolved_tickets', 0),
                            f"{tech.get('avg_resolution_hours', 0):.1f}"
                        ])

            logger.info(f"CSV report generated successfully: {file_path}")

        except Exception as e:
            logger.error(f"Error generating CSV report: {e}")
            raise

    def _generate_text_fallback(self, file_path: str, config: Dict, data: Dict, format_type: str):
        """Generate simple text file as fallback when libraries are not available"""
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"LANET Helpdesk - Reporte del Sistema ({format_type} Format)\n")
                f.write("=" * 60 + "\n\n")

                f.write(f"Reporte: {config.get('name', 'Sin nombre')}\n")
                f.write(f"Tipo: {config.get('report_type', 'dashboard').title()}\n")
                f.write(f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')}\n")
                f.write(f"Período: {data.get('period', {}).get('start_date', 'N/A')} - {data.get('period', {}).get('end_date', 'N/A')}\n\n")

                # Tickets data
                if 'tickets' in data:
                    f.write("MÉTRICAS DE TICKETS\n")
                    f.write("-" * 30 + "\n")
                    tickets_data = data['tickets']

                    f.write(f"Total de Tickets: {tickets_data.get('total_tickets', 0)}\n")
                    f.write(f"Tickets Abiertos: {tickets_data.get('open_tickets', 0)}\n")
                    f.write(f"Tickets Resueltos: {tickets_data.get('resolved_tickets', 0)}\n")
                    f.write(f"Tickets Cerrados: {tickets_data.get('closed_tickets', 0)}\n")
                    f.write(f"Prioridad Alta: {tickets_data.get('high_priority', 0)}\n")
                    f.write(f"Prioridad Media: {tickets_data.get('medium_priority', 0)}\n")
                    f.write(f"Prioridad Baja: {tickets_data.get('low_priority', 0)}\n")
                    f.write(f"Tiempo Promedio Resolución (hrs): {tickets_data.get('avg_resolution_hours', 0):.1f}\n\n")

                # SLA data
                if 'sla' in data:
                    f.write("MÉTRICAS DE SLA\n")
                    f.write("-" * 30 + "\n")
                    sla_data = data['sla']

                    f.write(f"Tickets con SLA: {sla_data.get('total_tickets_with_sla', 0)}\n")
                    f.write(f"Incumplimientos Respuesta: {sla_data.get('response_breaches', 0)}\n")
                    f.write(f"Incumplimientos Resolución: {sla_data.get('resolution_breaches', 0)}\n")
                    f.write(f"Cumplimiento Respuesta (%): {sla_data.get('response_compliance', 0):.1f}%\n")
                    f.write(f"Cumplimiento Resolución (%): {sla_data.get('resolution_compliance', 0):.1f}%\n")
                    f.write(f"Tiempo Promedio Respuesta (min): {sla_data.get('avg_response_minutes', 0):.1f}\n")
                    f.write(f"Tiempo Promedio Resolución (min): {sla_data.get('avg_resolution_minutes', 0):.1f}\n\n")

                # Technicians data
                if 'technicians' in data and data['technicians'].get('technicians'):
                    f.write("RENDIMIENTO DE TÉCNICOS\n")
                    f.write("-" * 30 + "\n")

                    for tech in data['technicians']['technicians']:
                        f.write(f"Técnico: {tech.get('first_name', '')} {tech.get('last_name', '')}\n")
                        f.write(f"  Tickets Asignados: {tech.get('assigned_tickets', 0)}\n")
                        f.write(f"  Tickets Resueltos: {tech.get('resolved_tickets', 0)}\n")
                        f.write(f"  Tiempo Promedio (hrs): {tech.get('avg_resolution_hours', 0):.1f}\n\n")

                f.write(f"\nNOTA: Este reporte fue generado en formato de texto porque las librerías de {format_type} no están disponibles.\n")
                f.write("Para reportes con formato profesional, instale las dependencias requeridas:\n")
                f.write("- PDF: pip install reportlab\n")
                f.write("- Excel: pip install openpyxl\n")

            logger.info(f"Text fallback report generated: {file_path}")

        except Exception as e:
            logger.error(f"Error generating text fallback report: {e}")
            raise


# Create generator instance
report_generator = ReportGenerator()
