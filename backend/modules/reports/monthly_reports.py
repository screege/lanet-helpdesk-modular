#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LANET Helpdesk V3 - Monthly Reports System
Simple automated monthly reports like ServiceDesk Plus
"""

import logging
import os
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import pytz
from flask import current_app

logger = logging.getLogger(__name__)

class MonthlyReportsService:
    """Simple service for automated monthly reports"""
    
    def __init__(self):
        self.mexico_tz = pytz.timezone('America/Mexico_City')
        self.reports_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports_files')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def get_system_status(self) -> Dict:
        """Get status of monthly reports system"""
        try:
            # Get next month first day
            now = datetime.now(self.mexico_tz)
            if now.month == 12:
                next_month = datetime(now.year + 1, 1, 1, 6, 0, 0)
            else:
                next_month = datetime(now.year, now.month + 1, 1, 6, 0, 0)
            next_month = self.mexico_tz.localize(next_month)
            
            # Count active clients
            clients_query = "SELECT COUNT(*) as count FROM clients WHERE is_active = true"
            client_count = current_app.db_manager.execute_query(clients_query, fetch='one')
            
            # Get last execution
            last_exec_query = """
                SELECT started_at FROM report_executions 
                WHERE execution_type = 'scheduled' 
                ORDER BY started_at DESC LIMIT 1
            """
            last_exec = current_app.db_manager.execute_query(last_exec_query, fetch='one')
            
            return {
                'next_report': next_month.strftime('%d %B %Y, %H:%M'),
                'active_clients': client_count['count'] if client_count else 0,
                'last_report': last_exec['started_at'].strftime('%d %B %Y') if last_exec else 'Nunca',
                'system_active': True
            }
            
        except Exception as e:
            logger.error(f"Error getting system status: {e}")
            return {
                'next_report': 'Error',
                'active_clients': 0,
                'last_report': 'Error',
                'system_active': False
            }
    
    def generate_monthly_report_for_client(self, client_id: str, month: int, year: int) -> Optional[str]:
        """Generate monthly report for specific client"""
        try:
            logger.info(f"🚀 Starting report generation for client {client_id}, month {month}, year {year}")

            # Get client info
            client_query = "SELECT name FROM clients WHERE client_id = %s"
            client = current_app.db_manager.execute_query(client_query, (client_id,), fetch='one')

            if not client:
                logger.error(f"❌ Client {client_id} not found")
                return None

            client_name = client['name']
            logger.info(f"✅ Generating report for client: {client_name}")
            
            # Get tickets for the month
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
            
            tickets_query = """
                SELECT
                    t.ticket_number,
                    t.subject,
                    t.status,
                    t.priority,
                    t.created_at,
                    t.resolved_at,
                    tr.resolution_notes,
                    u.name as technician_name,
                    s.name as site_name
                FROM tickets t
                LEFT JOIN users u ON t.assigned_to = u.user_id
                LEFT JOIN sites s ON t.site_id = s.site_id
                LEFT JOIN ticket_resolutions tr ON t.ticket_id = tr.ticket_id
                WHERE t.client_id = %s
                AND t.created_at >= %s
                AND t.created_at < %s
                ORDER BY t.created_at DESC
            """
            
            tickets = current_app.db_manager.execute_query(
                tickets_query,
                (client_id, start_date, end_date),
                fetch='all'
            )

            logger.info(f"📊 Found {len(tickets)} tickets for the period")

            # Generate Excel file
            filename = f"lanet_systems_informe_{client_name.lower().replace(' ', '_')}_{year}_{month:02d}.xlsx"
            file_path = os.path.join(self.reports_dir, filename)
            logger.info(f"📁 Target file path: {file_path}")
            
            # Create report file - Try Excel first, fallback to text
            try:
                # Try Excel first
                logger.info(f"Creating Excel report: {file_path}")
                self._create_simple_excel_report(client_name, tickets, file_path, month, year)

                if os.path.exists(file_path):
                    logger.info(f"Excel report file created successfully: {file_path}")
                    return file_path
                else:
                    logger.warning(f"Excel report file was not created, trying text fallback")
                    # Fallback to text
                    text_file_path = file_path.replace('.xlsx', '.txt')
                    logger.info(f"Creating text report: {text_file_path}")

                    self._create_text_report(client_name, tickets, text_file_path, month, year)

                    if os.path.exists(text_file_path):
                        logger.info(f"Text report file created successfully: {text_file_path}")
                        return text_file_path
                    else:
                        logger.error(f"Text report file was not created: {text_file_path}")
                        return None

            except Exception as e:
                logger.error(f"Error creating report: {e}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
                return None
            
        except Exception as e:
            logger.error(f"❌ Error generating monthly report for client {client_id}: {e}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            return None

    def _create_simple_excel_report(self, client_name: str, tickets: List[Dict], file_path: str, month: int, year: int):
        """Create a simple Excel report using openpyxl"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            logger.info(f"Creating simple Excel report for {client_name} with {len(tickets)} tickets")

            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Mensual"

            # Header with company branding
            ws['A1'] = "LANET SYSTEMS - REPORTE MENSUAL DE SOPORTE TÉCNICO"
            ws['A2'] = f"Cliente: {client_name.upper()}"
            ws['A3'] = f"Período: {month:02d}/{year}"
            ws['A4'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A5'] = f"Total de Tickets: {len(tickets)}"

            # Style header
            header_font = Font(bold=True, size=14)
            ws['A1'].font = header_font

            # Column headers
            headers = ['Número', 'Asunto', 'Estado', 'Prioridad', 'Creado', 'Resuelto', 'Técnico', 'Sitio']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Data rows
            for row, ticket in enumerate(tickets, 8):
                ws.cell(row=row, column=1, value=ticket.get('ticket_number', 'N/A'))
                ws.cell(row=row, column=2, value=ticket.get('subject', 'Sin asunto'))
                ws.cell(row=row, column=3, value=ticket.get('status', 'N/A'))
                ws.cell(row=row, column=4, value=ticket.get('priority', 'N/A'))

                # Format dates
                created_at = ticket.get('created_at')
                if created_at:
                    ws.cell(row=row, column=5, value=created_at.strftime('%d/%m/%Y %H:%M'))
                else:
                    ws.cell(row=row, column=5, value='N/A')

                resolved_at = ticket.get('resolved_at')
                if resolved_at:
                    ws.cell(row=row, column=6, value=resolved_at.strftime('%d/%m/%Y %H:%M'))
                else:
                    ws.cell(row=row, column=6, value='Pendiente')

                ws.cell(row=row, column=7, value=ticket.get('technician_name', 'Sin asignar'))
                ws.cell(row=row, column=8, value=ticket.get('site_name', 'Sin sitio'))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save file
            wb.save(file_path)
            logger.info(f"Simple Excel report saved to: {file_path}")

        except Exception as e:
            logger.error(f"Error creating simple Excel report: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def _clean_excel_value(self, value):
        """Clean value for Excel to avoid XML corruption"""
        if value is None:
            return ""

        # Convert to string
        str_value = str(value)

        # Remove or replace problematic characters
        str_value = str_value.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')

        # Remove control characters (ASCII 0-31 except tab, newline, carriage return)
        cleaned = ''.join(char for char in str_value if ord(char) >= 32 or char in '\t\n\r')

        # Limit length to prevent Excel issues
        if len(cleaned) > 32767:  # Excel cell limit
            cleaned = cleaned[:32767]

        return cleaned

    def _create_comprehensive_excel_report(self, tickets: List[Dict], file_path: str, start_date: datetime, end_date: datetime, report_title: str):
        """Create comprehensive Excel report with ALL tickets and client column - ULTRA SAFE VERSION"""
        try:
            import pandas as pd

            logger.info(f"Creating ULTRA SAFE Excel report with {len(tickets)} tickets using pandas")

            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Prepare data for pandas DataFrame
            data_rows = []

            for ticket in tickets:
                # Ultra safe data cleaning
                row = {
                    'Número': self._ultra_clean_value(ticket.get('ticket_number', 'N/A')),
                    'Cliente': self._ultra_clean_value(ticket.get('client_name', 'Sin cliente')),
                    'Asunto': self._ultra_clean_value(ticket.get('subject', 'Sin asunto')),
                    'Estado': self._ultra_clean_value(ticket.get('status', 'N/A')),
                    'Prioridad': self._ultra_clean_value(ticket.get('priority', 'N/A')),
                    'Creado': self._format_date_safe(ticket.get('created_at')),
                    'Resuelto': self._format_date_safe(ticket.get('resolved_at'), 'Pendiente'),
                    'Técnico': self._ultra_clean_value(ticket.get('technician_name', 'Sin asignar')),
                    'Sitio': self._ultra_clean_value(ticket.get('site_name', 'Sin sitio')),
                    'Solución': self._ultra_clean_value(ticket.get('resolution_notes', 'Pendiente'))
                }
                data_rows.append(row)

            # Create DataFrame
            df = pd.DataFrame(data_rows)

            # Create Excel writer with safe options
            with pd.ExcelWriter(file_path, engine='openpyxl', options={'remove_timezone': True}) as writer:
                # Write data starting from row 8 to leave space for headers
                df.to_excel(writer, sheet_name='Reporte Consolidado', index=False, startrow=7)

                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Reporte Consolidado']

                # Add custom headers
                worksheet['A1'] = "LANET SYSTEMS - REPORTE CONSOLIDADO"
                worksheet['A2'] = self._ultra_clean_value(report_title)
                worksheet['A3'] = f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
                worksheet['A4'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                worksheet['A5'] = f"Total de Tickets: {len(tickets)}"

                # Style headers (simple, no complex formatting)
                from openpyxl.styles import Font
                worksheet['A1'].font = Font(bold=True, size=14)
                worksheet['A2'].font = Font(bold=True, size=12)

                # Auto-adjust column widths (simple version)
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)  # Reduced max width
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info(f"ULTRA SAFE Excel report saved to: {file_path}")

        except Exception as e:
            logger.error(f"Error creating ULTRA SAFE Excel report: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def _ultra_clean_value(self, value):
        """Ultra aggressive cleaning for Excel compatibility"""
        if value is None:
            return ""

        # Convert to string
        str_value = str(value)

        # Remove ALL potentially problematic characters
        # Keep only basic alphanumeric, spaces, and common punctuation
        allowed_chars = set('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 .,;:!?()-_@#$%&*+=[]{}|\\/<>')
        cleaned = ''.join(char if char in allowed_chars else ' ' for char in str_value)

        # Remove multiple spaces
        cleaned = ' '.join(cleaned.split())

        # Limit length aggressively
        if len(cleaned) > 255:  # Much shorter limit
            cleaned = cleaned[:255]

        return cleaned

    def _format_date_safe(self, date_value, default='N/A'):
        """Safely format dates for Excel"""
        if date_value is None:
            return default

        try:
            if hasattr(date_value, 'strftime'):
                return date_value.strftime('%d/%m/%Y %H:%M')
            else:
                return str(date_value)
        except:
            return default

    def _create_csv_report(self, tickets: List[Dict], file_path: str, start_date: datetime, end_date: datetime, report_title: str):
        """Create CSV report - most compatible format"""
        try:
            import csv

            logger.info(f"Creating CSV report with {len(tickets)} tickets")

            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:  # utf-8-sig for Excel compatibility
                # Define headers
                headers = ['Número', 'Cliente', 'Asunto', 'Estado', 'Prioridad', 'Creado', 'Resuelto', 'Técnico', 'Sitio', 'Solución']

                writer = csv.DictWriter(csvfile, fieldnames=headers)

                # Write header info
                csvfile.write(f"LANET SYSTEMS - REPORTE CONSOLIDADO DE SOPORTE TÉCNICO\n")
                csvfile.write(f"{report_title}\n")
                csvfile.write(f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}\n")
                csvfile.write(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
                csvfile.write(f"Total de Tickets: {len(tickets)}\n")
                csvfile.write("\n")  # Empty line

                # Write column headers
                writer.writeheader()

                # Write data
                for ticket in tickets:
                    row = {
                        'Número': self._ultra_clean_value(ticket.get('ticket_number', 'N/A')),
                        'Cliente': self._ultra_clean_value(ticket.get('client_name', 'Sin cliente')),
                        'Asunto': self._ultra_clean_value(ticket.get('subject', 'Sin asunto')),
                        'Estado': self._ultra_clean_value(ticket.get('status', 'N/A')),
                        'Prioridad': self._ultra_clean_value(ticket.get('priority', 'N/A')),
                        'Creado': self._format_date_safe(ticket.get('created_at')),
                        'Resuelto': self._format_date_safe(ticket.get('resolved_at'), 'Pendiente'),
                        'Técnico': self._ultra_clean_value(ticket.get('technician_name', 'Sin asignar')),
                        'Sitio': self._ultra_clean_value(ticket.get('site_name', 'Sin sitio')),
                        'Solución': self._ultra_clean_value(ticket.get('resolution_notes', 'Pendiente'))
                    }
                    writer.writerow(row)

            logger.info(f"CSV report saved to: {file_path}")

        except Exception as e:
            logger.error(f"Error creating CSV report: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def _create_clean_excel_report(self, tickets: List[Dict], file_path: str, start_date: datetime, end_date: datetime, report_title: str):
        """Create professional Excel report with corruption protection"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            logger.info(f"Creating professional Excel report with {len(tickets)} tickets")

            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Consolidado"

            # Define professional styles
            title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')

            subtitle_font = Font(name='Arial', size=12, bold=True)
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

            data_font = Font(name='Arial', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title section - Updated to span 11 columns for new structure
            ws.merge_cells('A1:K1')
            title_cell = ws['A1']
            title_cell.value = 'LANET SYSTEMS - REPORTE CONSOLIDADO DE SOPORTE TÉCNICO'
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = center_alignment
            title_cell.border = thin_border

            # Subtitle and info
            ws.merge_cells('A2:K2')
            ws['A2'].value = report_title
            ws['A2'].font = subtitle_font
            ws['A2'].alignment = center_alignment

            ws.merge_cells('A3:K3')
            ws['A3'].value = f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            ws['A3'].font = data_font
            ws['A3'].alignment = center_alignment

            ws.merge_cells('A4:K4')
            ws['A4'].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total de Tickets: {len(tickets)}"
            ws['A4'].font = data_font
            ws['A4'].alignment = center_alignment

            # Column headers - Replace "Solución" with "Fecha Resolución" and "Resolución" (11 total)
            headers = ['Número', 'Cliente', 'Asunto', 'Estado', 'Prioridad', 'Creado', 'Resuelto', 'Técnico', 'Sitio', 'Fecha Resolución', 'Resolución']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # Data rows - Replace "Solución" with "Fecha Resolución" and "Resolución" (11 columns)
            for row_idx, ticket in enumerate(tickets, 7):
                # Apply corruption-safe cleaning while preserving readability
                row_data = [
                    self._safe_clean_string(ticket.get('ticket_number', 'N/A')),                    # 1. Número
                    self._safe_clean_string(ticket.get('client_name', 'Sin cliente')),             # 2. Cliente
                    self._safe_clean_string(ticket.get('subject', 'Sin asunto')),                  # 3. Asunto
                    self._safe_clean_string(ticket.get('status', 'N/A')),                          # 4. Estado
                    self._safe_clean_string(ticket.get('priority', 'N/A')),                        # 5. Prioridad
                    self._safe_date_format(ticket.get('created_at')),                              # 6. Creado
                    self._safe_date_format(ticket.get('resolved_at'), 'Pendiente'),                # 7. Resuelto
                    self._safe_clean_string(ticket.get('technician_name', 'Sin asignar')),         # 8. Técnico
                    self._safe_clean_string(ticket.get('site_name', 'Sin sitio')),                 # 9. Sitio
                    self._safe_date_format(ticket.get('resolved_at'), 'Sin fecha'),                # 10. Fecha Resolución
                    self._safe_clean_string(ticket.get('resolution_notes') or 'Sin resolución')    # 11. Resolución
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    if col in [1, 4, 5]:  # Center align ticket number, status, priority
                        cell.alignment = center_alignment

            # Set professional column widths - Updated for 11 columns
            column_widths = [12, 25, 35, 12, 10, 16, 16, 20, 20, 18, 35]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            # Set row heights
            ws.row_dimensions[1].height = 25
            for row in range(2, 5):
                ws.row_dimensions[row].height = 18
            ws.row_dimensions[6].height = 20

            # Save workbook
            wb.save(file_path)
            logger.info(f"Professional Excel report saved to: {file_path}")

        except Exception as e:
            logger.error(f"Error creating professional Excel report: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def _ultra_clean_string(self, value):
        """DEFINITIVE Excel-safe string cleaning - removes formula characters"""
        if value is None:
            return ""

        # Convert to string
        str_value = str(value).strip()

        if not str_value:
            return ""

        # REMOVE FORMULA CHARACTERS THAT CAUSE CORRUPTION
        # These are the exact characters causing the Excel XML corruption
        if str_value.startswith('='):
            str_value = str_value[1:]  # Remove leading =

        # Remove other formula starters
        if str_value.startswith(('+', '-', '@')):
            str_value = ' ' + str_value  # Add space to prevent formula interpretation

        # Keep ONLY basic ASCII letters, numbers, spaces, and safe punctuation
        # NO special characters that can corrupt XML
        allowed_chars = set('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 .,()-:')

        cleaned = ''.join(char if char in allowed_chars else ' ' for char in str_value)

        # Remove multiple spaces
        cleaned = ' '.join(cleaned.split())

        # Limit length aggressively to prevent Excel issues
        if len(cleaned) > 50:
            cleaned = cleaned[:50]

        # Final safety check - ensure no formula start
        if cleaned and cleaned[0] in '=+-@':
            cleaned = ' ' + cleaned

        return cleaned if cleaned else ""

    def _safe_date_format(self, date_value, default='N/A'):
        """Safe date formatting for Excel"""
        if date_value is None:
            return default

        try:
            if hasattr(date_value, 'strftime'):
                return date_value.strftime('%d/%m/%Y %H:%M')
            else:
                return str(date_value)
        except Exception:
            return default

    def _excel_safe_string(self, value):
        """Excel-safe string cleaning - prevents XML corruption"""
        if value is None:
            return ""

        # Convert to string and strip
        str_value = str(value).strip()

        if not str_value:
            return ""

        # Remove characters that cause Excel XML corruption
        # Keep only safe characters: letters, numbers, basic punctuation, spaces
        safe_chars = []
        for char in str_value:
            char_code = ord(char)
            # Allow basic ASCII printable characters and common Spanish characters
            if (32 <= char_code <= 126 or
                char in 'áéíóúñüÁÉÍÓÚÑÜ¿¡' or
                char_code in [160, 161, 191]):  # Non-breaking space, inverted punctuation
                safe_chars.append(char)
            elif char in '\t\n\r':
                safe_chars.append(' ')  # Replace line breaks with spaces
            # Skip all other characters that might cause XML issues

        # Join and clean up spaces
        cleaned = ''.join(safe_chars)
        cleaned = ' '.join(cleaned.split())  # Remove multiple spaces

        # Limit length to Excel's cell limit
        if len(cleaned) > 32767:
            cleaned = cleaned[:32767]

        return cleaned if cleaned else ""

    def _excel_safe_date(self, date_value, default='N/A'):
        """Excel-safe date formatting"""
        if date_value is None:
            return default

        try:
            if hasattr(date_value, 'strftime'):
                # Return simple date string - no complex formatting
                return date_value.strftime('%d/%m/%Y %H:%M')
            else:
                return str(date_value)
        except Exception:
            return default

    def _safe_clean_string(self, value):
        """Safe cleaning that prevents corruption while maintaining readability"""
        if value is None:
            return "Sin información"

        # Convert to string
        str_value = str(value).strip()

        if not str_value:
            return ""

        # Remove formula characters that cause corruption
        if str_value.startswith('='):
            str_value = str_value[1:]  # Remove leading =

        if str_value.startswith(('+', '-', '@')):
            str_value = ' ' + str_value  # Add space to prevent formula interpretation

        # Keep readable characters including Spanish accents
        safe_chars = []
        for char in str_value:
            char_code = ord(char)
            if (32 <= char_code <= 126 or
                char in 'áéíóúñüÁÉÍÓÚÑÜ¿¡'):
                safe_chars.append(char)
            elif char in '\t\n\r':
                safe_chars.append(' ')
            else:
                safe_chars.append(' ')  # Replace problematic chars with space

        # Join and clean up spaces
        cleaned = ''.join(safe_chars)
        cleaned = ' '.join(cleaned.split())

        # Reasonable length limit
        if len(cleaned) > 200:
            cleaned = cleaned[:200] + '...'

        # Final safety check
        if cleaned and cleaned[0] in '=+-@':
            cleaned = ' ' + cleaned

        return cleaned if cleaned else ""

    def _create_excel_report(self, client_name: str, tickets: List[Dict], file_path: str, month: int, year: int):
        """Create Excel report file"""
        try:
            logger.info(f"Creating Excel report for {client_name} with {len(tickets)} tickets")

            # Try pandas first
            try:
                import pandas as pd

                # Prepare data
                report_data = []
                for ticket in tickets:
                    report_data.append({
                        'Número de Ticket': ticket['ticket_number'],
                        'Asunto': ticket['subject'],
                        'Estado': ticket['status'],
                        'Prioridad': ticket['priority'],
                        'Fecha Creación': ticket['created_at'].strftime('%d/%m/%Y %H:%M') if ticket['created_at'] else '',
                        'Fecha Resolución': ticket['resolved_at'].strftime('%d/%m/%Y %H:%M') if ticket['resolved_at'] else 'Pendiente',
                        'Técnico Asignado': ticket['technician_name'] or 'Sin asignar',
                        'Sitio': ticket['site_name'] or 'Sin sitio'
                    })

                # Create DataFrame
                df = pd.DataFrame(report_data)

                # Create Excel file with multiple sheets
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Summary sheet
                    summary_data = {
                        'Métrica': [
                            'Cliente',
                            'Período',
                            'Total de Tickets',
                            'Tickets Resueltos',
                            'Tickets Pendientes',
                            'Tiempo Promedio de Resolución'
                        ],
                        'Valor': [
                            client_name,
                            f"{month:02d}/{year}",
                            len(tickets),
                            len([t for t in tickets if t['resolved_at']]),
                            len([t for t in tickets if not t['resolved_at']]),
                            'Calculando...'
                        ]
                    }

                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Resumen', index=False)

                    # Tickets detail sheet
                    if not df.empty:
                        df.to_excel(writer, sheet_name='Detalle de Tickets', index=False)
                    else:
                        # Empty sheet with headers
                        empty_df = pd.DataFrame(columns=[
                            'Número de Ticket', 'Asunto', 'Estado', 'Prioridad',
                            'Fecha Creación', 'Fecha Resolución', 'Técnico Asignado', 'Sitio'
                        ])
                        empty_df.to_excel(writer, sheet_name='Detalle de Tickets', index=False)

                logger.info(f"Excel report created successfully: {file_path}")

            except ImportError:
                logger.warning("Pandas not available, creating simple text report")
                raise Exception("Pandas not available")

        except Exception as e:
            logger.error(f"Error creating Excel report: {e}")
            # Fallback to simple text file
            text_file_path = file_path.replace('.xlsx', '.txt')
            with open(text_file_path, 'w', encoding='utf-8') as f:
                f.write(f"REPORTE MENSUAL - {client_name}\n")
                f.write(f"Período: {month:02d}/{year}\n")
                f.write(f"Total de tickets: {len(tickets)}\n\n")

                f.write("RESUMEN:\n")
                f.write(f"- Tickets Resueltos: {len([t for t in tickets if t['resolved_at']])}\n")
                f.write(f"- Tickets Pendientes: {len([t for t in tickets if not t['resolved_at']])}\n\n")

                f.write("DETALLE DE TICKETS:\n")
                for ticket in tickets:
                    f.write(f"- {ticket['ticket_number']}: {ticket['subject']} ({ticket['status']})\n")

            logger.info(f"Text report created as fallback: {text_file_path}")
            return text_file_path

    def _create_text_report(self, client_name: str, tickets: List[Dict], file_path: str, month: int, year: int):
        """Create simple text report file"""
        try:
            logger.info(f"Creating text report for {client_name} with {len(tickets)} tickets")

            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            logger.info(f"Creating text report at: {file_path}")

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("=" * 70 + "\n")
                f.write("LANET SYSTEMS - REPORTE MENSUAL DE SOPORTE TÉCNICO\n")
                f.write("=" * 70 + "\n\n")

                f.write(f"Cliente: {client_name.upper()}\n")
                f.write(f"Período: {month:02d}/{year}\n")
                f.write(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n")

                # Summary section
                f.write("RESUMEN EJECUTIVO\n")
                f.write("-" * 20 + "\n")
                f.write(f"Total de tickets: {len(tickets)}\n")
                f.write(f"Tickets resueltos: {len([t for t in tickets if t['resolved_at']])}\n")
                f.write(f"Tickets pendientes: {len([t for t in tickets if not t['resolved_at']])}\n\n")

                # Tickets by priority
                priorities = {}
                for ticket in tickets:
                    priority = ticket['priority'] or 'Sin prioridad'
                    priorities[priority] = priorities.get(priority, 0) + 1

                f.write("TICKETS POR PRIORIDAD\n")
                f.write("-" * 20 + "\n")
                for priority, count in priorities.items():
                    f.write(f"{priority}: {count}\n")
                f.write("\n")

                # Tickets by status
                statuses = {}
                for ticket in tickets:
                    status = ticket['status'] or 'Sin estado'
                    statuses[status] = statuses.get(status, 0) + 1

                f.write("TICKETS POR ESTADO\n")
                f.write("-" * 18 + "\n")
                for status, count in statuses.items():
                    f.write(f"{status}: {count}\n")
                f.write("\n")

                # Detailed tickets list
                f.write("DETALLE DE TICKETS\n")
                f.write("-" * 18 + "\n")

                if tickets:
                    for i, ticket in enumerate(tickets, 1):
                        f.write(f"{i}. {ticket['ticket_number']}\n")
                        f.write(f"   Asunto: {ticket['subject']}\n")
                        f.write(f"   Estado: {ticket['status']}\n")
                        f.write(f"   Prioridad: {ticket['priority'] or 'Sin prioridad'}\n")
                        f.write(f"   Creado: {ticket['created_at'].strftime('%d/%m/%Y %H:%M') if ticket['created_at'] else 'N/A'}\n")
                        f.write(f"   Resuelto: {ticket['resolved_at'].strftime('%d/%m/%Y %H:%M') if ticket['resolved_at'] else 'Pendiente'}\n")
                        f.write(f"   Técnico: {ticket['technician_name'] or 'Sin asignar'}\n")
                        f.write(f"   Sitio: {ticket['site_name'] or 'Sin sitio'}\n")
                        f.write("\n")
                else:
                    f.write("No se encontraron tickets para este período.\n\n")

                f.write("=" * 60 + "\n")
                f.write("Fin del reporte\n")
                f.write("=" * 60 + "\n")

            logger.info(f"Text report created successfully: {file_path}")

        except Exception as e:
            logger.error(f"Error creating text report: {e}")
            raise
    
    def generate_test_report(self, user_claims: Dict) -> Optional[str]:
        """Generate a comprehensive test report with ALL tickets from ALL clients"""
        try:
            user_role = user_claims.get('role')
            logger.info(f"🧪 Generating comprehensive test report for user role: {user_role}")

            # For superadmin/technician: Generate report with ALL tickets from ALL clients
            if user_role in ['superadmin', 'technician']:
                return self.generate_comprehensive_report()

            # For client_admin: Generate only their client's tickets
            elif user_role == 'client_admin':
                client_id = user_claims.get('client_id')
                if client_id:
                    now = datetime.now()
                    return self.generate_monthly_report_for_client(client_id, now.month, now.year)
                else:
                    logger.error("❌ Client admin without client_id")
                    return None
            else:
                logger.error(f"❌ Invalid user role for test report: {user_role}")
                return None

        except Exception as e:
            logger.error(f"❌ Error generating test report: {e}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            return None

    def generate_comprehensive_report(self, start_date=None, end_date=None, client_id=None) -> Optional[str]:
        """Generate comprehensive report with ALL tickets from ALL clients"""
        try:
            logger.info(f"🚀 Generating comprehensive report...")
            logger.info(f"📋 Parameters: start_date={start_date}, end_date={end_date}, client_id={client_id}")

            # Default to current month if no dates provided
            if not start_date or not end_date:
                now = datetime.now()
                start_date = datetime(now.year, now.month, 1)
                if now.month == 12:
                    end_date = datetime(now.year + 1, 1, 1)
                else:
                    end_date = datetime(now.year, now.month + 1, 1)

            # Ensure end_date includes the full day
            if hasattr(end_date, 'hour') and end_date.hour == 0:
                end_date = end_date.replace(hour=23, minute=59, second=59)

            logger.info(f"📅 Final date range: {start_date} to {end_date}")

            # Build query based on filters
            where_conditions = ["t.created_at >= %s", "t.created_at < %s"]
            params = [start_date, end_date]

            if client_id:
                where_conditions.append("t.client_id = %s")
                params.append(client_id)
                report_title = "REPORTE POR CLIENTE"
            else:
                report_title = "REPORTE CONSOLIDADO - TODOS LOS CLIENTES"

            # Query to get ALL tickets with client information and resolution from ticket_resolutions table
            tickets_query = f"""
                SELECT
                    t.ticket_number,
                    t.subject,
                    t.status,
                    t.priority,
                    t.created_at,
                    t.resolved_at,
                    tr.resolution_notes,
                    u.name as technician_name,
                    s.name as site_name,
                    c.name as client_name
                FROM tickets t
                LEFT JOIN users u ON t.assigned_to = u.user_id
                LEFT JOIN sites s ON t.site_id = s.site_id
                LEFT JOIN clients c ON t.client_id = c.client_id
                LEFT JOIN ticket_resolutions tr ON t.ticket_id = tr.ticket_id
                WHERE {' AND '.join(where_conditions)}
                ORDER BY t.created_at DESC
            """

            tickets = current_app.db_manager.execute_query(tickets_query, params, fetch='all')
            logger.info(f"📊 Found {len(tickets)} tickets total")

            # Generate filename - BACK TO EXCEL FORMAT AS REQUESTED
            date_str = start_date.strftime('%Y_%m')
            if client_id:
                filename = f"reporte_cliente_{date_str}.xlsx"
            else:
                filename = f"reporte_consolidado_{date_str}.xlsx"

            file_path = os.path.join(self.reports_dir, filename)

            # Create CLEAN Excel report using proven method
            try:
                logger.info(f"📄 Creating CLEAN Excel file: {file_path}")
                self._create_clean_excel_report(tickets, file_path, start_date, end_date, report_title)

                if os.path.exists(file_path):
                    file_size = os.path.getsize(file_path)
                    logger.info(f"✅ Simple Excel report created: {file_path} ({file_size} bytes)")
                    return file_path
                else:
                    logger.error(f"❌ Failed to create Excel report file")
                    return None

            except Exception as excel_error:
                logger.error(f"❌ Error creating Excel file: {excel_error}")
                import traceback
                logger.error(f"❌ Excel creation traceback: {traceback.format_exc()}")
                return None

        except Exception as e:
            logger.error(f"❌ Error generating comprehensive report: {e}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            return None
    
    def setup_monthly_schedules(self):
        """Setup automatic monthly schedules for all active clients"""
        try:
            # Get all active clients
            clients_query = "SELECT client_id, name FROM clients WHERE is_active = true"
            clients = current_app.db_manager.execute_query(clients_query, fetch='all')
            
            if not clients:
                logger.info("No active clients found")
                return
            
            # Create schedules for each client
            for client in clients:
                self._create_monthly_schedule_for_client(client['client_id'], client['name'])
            
            # Create consolidated schedule for superadmin
            self._create_consolidated_schedule()
            
            logger.info(f"Monthly schedules setup for {len(clients)} clients")
            
        except Exception as e:
            logger.error(f"Error setting up monthly schedules: {e}")
    
    def _create_monthly_schedule_for_client(self, client_id: str, client_name: str):
        """Create monthly schedule for specific client"""
        try:
            # Check if schedule already exists
            existing_query = """
                SELECT schedule_id FROM report_schedules rs
                JOIN report_configurations rc ON rs.config_id = rc.config_id
                WHERE rc.client_id = %s AND rs.schedule_type = 'monthly'
            """
            existing = current_app.db_manager.execute_query(existing_query, (client_id,), fetch='one')
            
            if existing:
                logger.info(f"Monthly schedule already exists for client {client_name}")
                return
            
            # Create configuration
            config_id = str(uuid.uuid4())
            config_query = """
                INSERT INTO report_configurations 
                (config_id, name, description, client_id, report_filters, output_formats, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            
            current_app.db_manager.execute_query(config_query, (
                config_id,
                f"Reporte Mensual - {client_name}",
                f"Reporte mensual automático para {client_name}",
                client_id,
                '{"type": "monthly"}',
                ['excel'],
                None  # System created
            ))
            
            # Create schedule
            schedule_id = str(uuid.uuid4())
            schedule_query = """
                INSERT INTO report_schedules 
                (schedule_id, name, config_id, schedule_type, schedule_config, recipients, next_run_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            
            # Calculate next run (first day of next month at 6 AM Mexico time)
            now = datetime.now(self.mexico_tz)
            if now.month == 12:
                next_run = datetime(now.year + 1, 1, 1, 6, 0, 0)
            else:
                next_run = datetime(now.year, now.month + 1, 1, 6, 0, 0)
            next_run = self.mexico_tz.localize(next_run)
            
            current_app.db_manager.execute_query(schedule_query, (
                schedule_id,
                f"Programación Mensual - {client_name}",
                config_id,
                'monthly',
                '{"day": 1, "hour": 6, "timezone": "America/Mexico_City"}',
                [],  # Recipients will be added later
                next_run
            ))
            
            logger.info(f"Created monthly schedule for {client_name}")
            
        except Exception as e:
            logger.error(f"Error creating monthly schedule for {client_name}: {e}")
    
    def _create_consolidated_schedule(self):
        """Create consolidated monthly schedule for superadmin"""
        try:
            # Check if consolidated schedule exists
            existing_query = """
                SELECT schedule_id FROM report_schedules 
                WHERE name LIKE '%Consolidado%' AND schedule_type = 'monthly'
            """
            existing = current_app.db_manager.execute_query(existing_query, fetch='one')
            
            if existing:
                logger.info("Consolidated schedule already exists")
                return
            
            # Create configuration
            config_id = str(uuid.uuid4())
            config_query = """
                INSERT INTO report_configurations 
                (config_id, name, description, client_id, report_filters, output_formats, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            
            current_app.db_manager.execute_query(config_query, (
                config_id,
                "Reporte Consolidado Mensual",
                "Reporte consolidado de todos los clientes para SuperAdmin",
                None,  # Global report
                '{"type": "consolidated"}',
                ['excel'],
                None  # System created
            ))
            
            # Create schedule
            schedule_id = str(uuid.uuid4())
            schedule_query = """
                INSERT INTO report_schedules 
                (schedule_id, name, config_id, schedule_type, schedule_config, recipients, next_run_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            
            # Calculate next run (first day of next month at 7 AM Mexico time)
            now = datetime.now(self.mexico_tz)
            if now.month == 12:
                next_run = datetime(now.year + 1, 1, 1, 7, 0, 0)
            else:
                next_run = datetime(now.year, now.month + 1, 1, 7, 0, 0)
            next_run = self.mexico_tz.localize(next_run)
            
            current_app.db_manager.execute_query(schedule_query, (
                schedule_id,
                "Programación Consolidada Mensual",
                config_id,
                'monthly',
                '{"day": 1, "hour": 7, "timezone": "America/Mexico_City"}',
                [],  # Recipients will be added later
                next_run
            ))
            
            logger.info("Created consolidated monthly schedule")
            
        except Exception as e:
            logger.error(f"Error creating consolidated schedule: {e}")

# Global instance
monthly_reports_service = MonthlyReportsService()
