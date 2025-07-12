#!/usr/bin/env python3
"""
Test Excel file structure in detail
"""

from openpyxl import load_workbook
import os

def test_excel_structure():
    """Test the Excel file structure in detail"""
    print("🔍 TESTING EXCEL STRUCTURE IN DETAIL")
    print("=" * 60)
    
    file_path = r"C:\lanet-helpdesk-v3\backend\reports_files\reporte_consolidado_2025_07.xlsx"
    
    try:
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return
        
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"📊 Sheet: {ws.title}")
        print(f"📊 Max row: {ws.max_row}")
        print(f"📊 Max column: {ws.max_column}")
        
        # Check headers (row 6)
        print(f"\n📋 Headers (Row 6):")
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=6, column=col).value
            headers.append(cell_value)
            print(f"  Column {col}: {repr(cell_value)}")
        
        print(f"\nTotal headers: {len(headers)}")
        
        # Check first data row (row 7)
        print(f"\n📋 First data row (Row 7):")
        first_row = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=7, column=col).value
            first_row.append(cell_value)
            print(f"  Column {col}: {repr(cell_value)}")
        
        print(f"\nTotal data columns: {len(first_row)}")
        
        # Check specific ticket #140 data
        print(f"\n🔍 Looking for ticket TKT-000140...")
        found_ticket = False
        for row in range(7, min(ws.max_row + 1, 20)):  # Check first 13 data rows
            ticket_number = ws.cell(row=row, column=1).value
            if ticket_number == 'TKT-000140':
                found_ticket = True
                print(f"✅ Found TKT-000140 in row {row}:")
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    header = headers[col-1] if col-1 < len(headers) else f"Col{col}"
                    print(f"  {header}: {repr(cell_value)}")
                break

        if not found_ticket:
            print("❌ TKT-000140 not found in first 13 rows")

        # Check for tickets with resolution data
        print(f"\n🔍 Looking for tickets with resolution data...")
        found_with_resolution = 0
        for row in range(7, min(ws.max_row + 1, 20)):  # Check first 13 data rows
            ticket_number = ws.cell(row=row, column=1).value
            resolution = ws.cell(row=row, column=11).value  # Column 11 is Resolución
            if resolution and resolution != 'Sin resolución':
                found_with_resolution += 1
                print(f"✅ {ticket_number}: {repr(resolution)}")
                if found_with_resolution >= 3:  # Show first 3
                    break

        if found_with_resolution == 0:
            print("❌ No tickets found with resolution data in first 13 rows")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    test_excel_structure()
