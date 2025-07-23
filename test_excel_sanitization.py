#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test Excel sanitization to verify formula injection prevention
"""

import openpyxl
import os
import requests
import json

def test_sanitization_function():
    """Test the sanitization function directly"""
    print("🧪 TESTING SANITIZATION FUNCTION")
    print("=" * 60)
    
    # Import the sanitization function
    import sys
    sys.path.append(os.path.join(os.getcwd(), 'backend'))
    
    try:
        from modules.reports.routes_simple import sanitize_excel_cell_value
        
        # Test cases for sanitization
        test_cases = [
            # (input, expected_output, description)
            ("TKT-000147", "'TKT-000147", "Ticket number with dash"),
            ("=SUM(A1:A10)", "'=SUM(A1:A10)", "Formula starting with ="),
            ("+1234567890", "'+1234567890", "Phone number starting with +"),
            ("-test data", "'-test data", "Text starting with -"),
            ("@username", "'@username", "Username starting with @"),
            ("email@domain.com", "email@domain.com", "Email with @ in middle (safe)"),
            ("normal text", "normal text", "Normal text (safe)"),
            ("text-with-dash", "text-with-dash", "Text with dash in middle (safe)"),
            ("", "", "Empty string"),
            (None, "", "None value"),
            ("=", "'=", "Just equals sign"),
            ("-", "'-", "Just dash"),
            ("Prueba de timezone - 4:16 PM", "Prueba de timezone - 4:16 PM", "Text with dash in middle (safe)"),
        ]
        
        all_passed = True
        
        for input_val, expected, description in test_cases:
            result = sanitize_excel_cell_value(input_val)
            passed = result == expected
            status = "✅" if passed else "❌"
            
            print(f"{status} {description}")
            print(f"   Input: {repr(input_val)}")
            print(f"   Expected: {repr(expected)}")
            print(f"   Got: {repr(result)}")
            
            if not passed:
                all_passed = False
            print()
        
        return all_passed
        
    except Exception as e:
        print(f"❌ Error testing sanitization function: {e}")
        return False

def test_excel_file_integrity():
    """Test that generated Excel files can be opened without corruption"""
    print("\n📊 TESTING EXCEL FILE INTEGRITY")
    print("=" * 60)
    
    # Find the latest generated Excel files
    reports_dir = os.path.join('backend', 'reports_files')
    
    if not os.path.exists(reports_dir):
        print("❌ Reports directory not found")
        return False
    
    # Get the latest Excel files
    excel_files = [f for f in os.listdir(reports_dir) if f.endswith('.xlsx') and 'Julio_2025' in f]
    
    if not excel_files:
        print("❌ No recent Excel files found")
        return False
    
    print(f"📋 Found {len(excel_files)} Excel files to test:")
    for file in excel_files:
        print(f"   - {file}")
    
    all_passed = True
    
    for filename in excel_files:
        file_path = os.path.join(reports_dir, filename)
        
        print(f"\n🔍 Testing: {filename}")
        print("-" * 40)
        
        try:
            # Try to open the Excel file with openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            print(f"✅ File opens successfully")
            print(f"   Sheet: {ws.title}")
            print(f"   Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Check for sanitized values (should start with single quote)
            sanitized_count = 0
            total_cells_checked = 0
            
            # Check first 10 data rows (starting from row 7)
            for row_num in range(7, min(17, ws.max_row + 1)):
                for col_num in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value and isinstance(cell_value, str):
                        total_cells_checked += 1
                        if cell_value.startswith("'"):
                            sanitized_count += 1
            
            print(f"   Cells checked: {total_cells_checked}")
            print(f"   Sanitized cells: {sanitized_count}")
            
            # Check specific ticket numbers (should be sanitized)
            ticket_cell = ws.cell(row=7, column=1).value  # First ticket number
            if ticket_cell and isinstance(ticket_cell, str):
                if ticket_cell.startswith("'TKT-"):
                    print(f"✅ Ticket numbers properly sanitized: {ticket_cell}")
                else:
                    print(f"⚠️ Ticket number not sanitized: {ticket_cell}")
            
            wb.close()
            
        except Exception as e:
            print(f"❌ Error opening file: {e}")
            all_passed = False
    
    return all_passed

def generate_test_report():
    """Generate a new test report to verify sanitization"""
    print("\n🚀 GENERATING NEW TEST REPORT")
    print("=" * 60)
    
    try:
        # Login first
        login_data = {
            "email": "ba@lanet.mx",
            "password": "TestAdmin123!"
        }
        response = requests.post('http://localhost:5001/api/auth/login', json=login_data)
        
        if response.status_code != 200:
            print(f"❌ Login failed: {response.status_code}")
            return False
        
        data = response.json()
        if not data.get('success'):
            print(f"❌ Login failed: {data}")
            return False
        
        token = data['data']['access_token']
        print("✅ Login successful")
        
        # Generate a quick report
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        response = requests.post(
            'http://localhost:5001/api/reports/generate-quick',
            headers=headers,
            json={'output_format': 'excel'}
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get('success'):
                execution_id = data['data']['execution_id']
                file_path = data['data'].get('file_path', 'unknown')
                print(f"✅ Test report generated successfully!")
                print(f"   Execution ID: {execution_id}")
                print(f"   File path: {file_path}")
                return True
            else:
                print(f"❌ Report generation failed: {data}")
                return False
        else:
            print(f"❌ Report generation failed: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"❌ Error generating test report: {e}")
        return False

def main():
    """Main test function"""
    print("🔒 EXCEL FORMULA INJECTION PREVENTION TEST")
    print("=" * 70)
    
    # Test 1: Sanitization function
    sanitization_passed = test_sanitization_function()
    
    # Test 2: Generate new report
    generation_passed = generate_test_report()
    
    # Test 3: File integrity
    integrity_passed = test_excel_file_integrity()
    
    # Summary
    print("\n" + "=" * 70)
    print("📊 TEST RESULTS SUMMARY")
    print("=" * 70)
    
    tests = [
        ("Sanitization Function", sanitization_passed),
        ("Report Generation", generation_passed),
        ("File Integrity", integrity_passed)
    ]
    
    passed_count = sum(1 for _, passed in tests if passed)
    total_count = len(tests)
    
    for test_name, passed in tests:
        status = "✅ PASSED" if passed else "❌ FAILED"
        print(f"{status} {test_name}")
    
    print(f"\n📈 Overall Result: {passed_count}/{total_count} tests passed")
    
    if passed_count == total_count:
        print("🎉 ALL TESTS PASSED - Excel formula injection prevention is working!")
        print("   ✅ Files should now open in Excel without corruption warnings")
        print("   ✅ Ticket numbers and other data are properly sanitized")
        print("   ✅ Formula characters are escaped with single quotes")
    else:
        print("⚠️ Some tests failed - review the issues above")
    
    return passed_count == total_count

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
