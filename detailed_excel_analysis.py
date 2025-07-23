#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Detailed analysis of the reference Excel file to understand the exact structure
"""

import openpyxl
import os

def detailed_analysis():
    """Detailed analysis of the reference Excel file"""
    file_path = r"C:\lanet-helpdesk-v3\reporte_consolidado_2025_07 (14).xlsx"
    
    if not os.path.exists(file_path):
        print(f"❌ Reference file not found: {file_path}")
        return
    
    try:
        print("📊 DETAILED ANALYSIS OF REFERENCE EXCEL FILE")
        print("=" * 70)
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"📋 Active sheet: {ws.title}")
        print(f"📐 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Analyze the first 20 rows to understand the structure
        print(f"\n📄 FIRST 20 ROWS ANALYSIS:")
        print("-" * 50)
        
        for row_num in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = cell.value
                if value is not None:
                    # Truncate long values but show more detail
                    str_value = str(value)
                    if len(str_value) > 40:
                        str_value = str_value[:40] + "..."
                    row_data.append(str_value)
                else:
                    row_data.append("None")
            
            print(f"Row {row_num:2d}: {row_data}")
        
        # Find where the actual data table starts
        print(f"\n🔍 LOOKING FOR DATA TABLE HEADERS:")
        print("-" * 50)
        
        data_start_row = None
        headers = []
        
        for row_num in range(1, min(30, ws.max_row + 1)):
            row_values = []
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    row_values.append(str(cell_value).strip())
            
            # Look for typical ticket headers
            row_text = " ".join(row_values).lower()
            if any(keyword in row_text for keyword in ['ticket', 'número', 'cliente', 'fecha', 'estado', 'prioridad']):
                print(f"📍 Potential header row {row_num}: {row_values}")
                if not data_start_row:
                    data_start_row = row_num
                    headers = row_values
        
        if data_start_row:
            print(f"\n✅ Data table starts at row {data_start_row}")
            print(f"📋 Headers: {headers}")
            
            # Show sample data rows
            print(f"\n📊 SAMPLE DATA ROWS:")
            print("-" * 50)
            
            for row_num in range(data_start_row + 1, min(data_start_row + 6, ws.max_row + 1)):
                row_data = []
                for col_num in range(1, len(headers) + 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value is not None:
                        str_value = str(cell_value)
                        if len(str_value) > 30:
                            str_value = str_value[:30] + "..."
                        row_data.append(str_value)
                    else:
                        row_data.append("None")
                
                print(f"Row {row_num:2d}: {row_data}")
            
            # Count actual data rows
            data_rows = 0
            for row_num in range(data_start_row + 1, ws.max_row + 1):
                has_data = False
                for col_num in range(1, len(headers) + 1):
                    if ws.cell(row=row_num, column=col_num).value:
                        has_data = True
                        break
                if has_data:
                    data_rows += 1
            
            print(f"\n📈 Total data rows: {data_rows}")
        
        else:
            print("❌ Could not identify data table headers")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")

if __name__ == "__main__":
    detailed_analysis()
