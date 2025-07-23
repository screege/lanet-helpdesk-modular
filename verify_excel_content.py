#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verify the content of the newly generated Excel files
"""

import openpyxl
import os

def verify_excel_file(filename):
    """Verify the content of an Excel file"""
    file_path = os.path.join('backend', 'reports_files', filename)
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        print(f"\n📊 ANALYZING: {filename}")
        print("=" * 60)
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"📋 Sheet name: {ws.title}")
        print(f"📐 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Show header information (first 5 rows)
        print(f"\n📄 HEADER SECTION:")
        print("-" * 40)
        for row_num in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    str_value = str(cell_value)
                    if len(str_value) > 50:
                        str_value = str_value[:50] + "..."
                    row_data.append(str_value)
                else:
                    row_data.append("None")
            print(f"Row {row_num}: {row_data}")
        
        # Find data headers (should be row 6)
        print(f"\n📋 DATA HEADERS (Row 6):")
        print("-" * 40)
        headers = []
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=6, column=col_num).value
            if cell_value:
                headers.append(str(cell_value))
        print(f"Headers: {headers}")
        
        # Count data rows (starting from row 7)
        data_rows = 0
        sample_data = []
        for row_num in range(7, ws.max_row + 1):
            has_data = False
            row_data = []
            for col_num in range(1, len(headers) + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    has_data = True
                    str_value = str(cell_value)
                    if len(str_value) > 20:
                        str_value = str_value[:20] + "..."
                    row_data.append(str_value)
                else:
                    row_data.append("None")
            
            if has_data:
                data_rows += 1
                if len(sample_data) < 3:  # Keep first 3 data rows as samples
                    sample_data.append(row_data)
        
        print(f"\n📊 DATA ANALYSIS:")
        print("-" * 40)
        print(f"Total data rows: {data_rows}")
        
        if sample_data:
            print(f"\nSample data rows:")
            for i, row in enumerate(sample_data, 1):
                print(f"  Row {i}: {row}")
        else:
            print("❌ No data rows found!")
        
        # Check for ticket numbers
        ticket_count = 0
        for row_num in range(7, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=1).value
            if cell_value and str(cell_value).startswith('TKT-'):
                ticket_count += 1
        
        print(f"\n🎫 Tickets found: {ticket_count}")
        
        wb.close()
        
        # Summary
        if data_rows > 0 and ticket_count > 0:
            print(f"✅ File contains {data_rows} data rows with {ticket_count} tickets")
        elif data_rows > 0:
            print(f"⚠️ File contains {data_rows} data rows but no ticket numbers found")
        else:
            print(f"❌ File appears to be empty (no data rows)")
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")

def main():
    """Main verification function"""
    print("🔍 VERIFYING EXCEL FILE CONTENT")
    print("=" * 70)
    
    # Check the new files with friendly names
    files_to_check = [
        'Reporte_Rapido_22_Julio_2025.xlsx',
        'Reporte_Mensual_Julio_2025.xlsx'
    ]
    
    for filename in files_to_check:
        verify_excel_file(filename)
    
    print("\n" + "=" * 70)
    print("📋 VERIFICATION COMPLETE")

if __name__ == "__main__":
    main()
