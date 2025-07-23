#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyze the reference Excel file to understand the expected structure and data format
"""

import openpyxl
import os

def analyze_reference_file():
    """Analyze the reference Excel file structure"""
    file_path = r"C:\lanet-helpdesk-v3\reporte_consolidado_2025_07 (14).xlsx"
    
    if not os.path.exists(file_path):
        print(f"❌ Reference file not found: {file_path}")
        return
    
    try:
        print("📊 Analyzing reference Excel file...")
        print("=" * 60)
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        print(f"📋 Workbook sheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n📄 Sheet: {sheet_name}")
            print("-" * 40)
            
            ws = wb[sheet_name]
            
            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"   Dimensions: {max_row} rows x {max_col} columns")
            
            # Get headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
            
            print(f"   Headers ({len(headers)}): {headers}")
            
            # Sample first few data rows
            print(f"   Sample data (first 3 rows):")
            for row in range(2, min(6, max_row + 1)):  # Skip header, show first 3 data rows
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:30])  # Truncate long values
                    else:
                        row_data.append("None")
                print(f"     Row {row}: {row_data}")
            
            # Count non-empty rows
            non_empty_rows = 0
            for row in range(2, max_row + 1):
                has_data = False
                for col in range(1, max_col + 1):
                    if ws.cell(row=row, column=col).value:
                        has_data = True
                        break
                if has_data:
                    non_empty_rows += 1
            
            print(f"   Non-empty data rows: {non_empty_rows}")
        
        wb.close()
        
        print("\n" + "=" * 60)
        print("✅ Analysis complete!")
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")

def check_current_database_data():
    """Check what data is currently in the database"""
    print("\n🔍 Checking current database data...")
    print("=" * 60)
    
    try:
        import sys
        import os
        sys.path.append(os.path.join(os.getcwd(), 'backend'))
        
        from core.database import DatabaseManager
        
        db_manager = DatabaseManager('postgresql://postgres:Poikl55+*@localhost:5432/lanet_helpdesk')
        
        # Check tickets count
        tickets_query = "SELECT COUNT(*) as count FROM tickets"
        result = db_manager.execute_query(tickets_query, fetch='one')
        print(f"📊 Total tickets in database: {result['count'] if result else 0}")
        
        # Check clients count
        clients_query = "SELECT COUNT(*) as count FROM clients WHERE is_active = true"
        result = db_manager.execute_query(clients_query, fetch='one')
        print(f"🏢 Active clients in database: {result['count'] if result else 0}")
        
        # Check sites count
        sites_query = "SELECT COUNT(*) as count FROM sites WHERE is_active = true"
        result = db_manager.execute_query(sites_query, fetch='one')
        print(f"📍 Active sites in database: {result['count'] if result else 0}")
        
        # Sample tickets data
        sample_query = """
        SELECT
            t.ticket_number,
            c.name as client_name,
            s.name as site_name,
            t.subject,
            t.status,
            t.priority,
            t.created_at
        FROM tickets t
        LEFT JOIN clients c ON t.client_id = c.client_id
        LEFT JOIN sites s ON t.site_id = s.site_id
        ORDER BY t.created_at DESC
        LIMIT 5
        """
        
        tickets = db_manager.execute_query(sample_query)
        
        print(f"\n📋 Sample tickets data:")
        if tickets:
            for ticket in tickets:
                print(f"   {ticket['ticket_number']}: {ticket['subject'][:50] if ticket['subject'] else 'No subject'}... ({ticket['status']})")
        else:
            print("   No tickets found")
        
    except Exception as e:
        print(f"❌ Error checking database: {e}")

if __name__ == "__main__":
    analyze_reference_file()
    check_current_database_data()
